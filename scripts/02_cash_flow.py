#!/usr/bin/env python3
"""
Cash Flow Report Generator
Generates weekly cash flow projections based on PO data and project timelines
"""
import sys
sys.path.append('/opt/torelo-kpi/scripts')
from config import DATABASE_CONFIG, BASE_DIR, WEB_DIR
import pandas as pd
from sqlalchemy import create_engine
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import os
import json

def get_cash_flow_data():
    """Get purchase orders with project task dates for cash flow analysis"""
    
    query = """
    WITH task_dates AS (
        SELECT 
            pt.id AS task_id,
            pt.name AS task_name,
            pt.fecha_inicio AS start_date,
            pt.date_deadline AS end_date,
            pp.name ->> 'es_GT' AS project_name,
            pp.id AS project_id
        FROM project_task pt
        JOIN project_project pp ON pt.project_id = pp.id
        WHERE pt.active = true
        AND pp.active = true
    ),
    po_tasks AS (
        SELECT 
            po.id AS po_id,
            po.name AS po_number,
            task_po.task_id,
            td.task_name,
            td.start_date,
            td.end_date,
            td.project_name,
            td.project_id
        FROM purchase_order po
        JOIN project_task_purchase_order_rel task_po ON po.id = task_po.purchase_order_id
        LEFT JOIN task_dates td ON task_po.task_id = td.task_id
        WHERE po.state NOT IN ('cancel', 'done')
    ),
    invoice_data AS (
        SELECT 
            pol.order_id,
            SUM(CASE 
                WHEN am.state != 'cancel' AND am.payment_state != 'paid' 
                THEN am.amount_residual_signed 
                ELSE 0 
            END) AS unpaid_amount
        FROM account_move am
        JOIN account_move_line aml ON aml.move_id = am.id
        JOIN purchase_order_line pol ON pol.id = aml.purchase_line_id
        WHERE am.move_type = 'in_invoice'
        GROUP BY pol.order_id
    )
    SELECT 
        po.id AS po_id,
        po.name AS po_number,
        po.state AS po_state,
        CASE 
            WHEN po.state = 'draft' THEN 'Borrador (RFQ)'
            WHEN po.state = 'sent' THEN 'Enviado'
            WHEN po.state = 'to approve' THEN 'Por Aprobar'
            WHEN po.state = 'purchase' THEN 'Orden de Compra'
            ELSE po.state
        END AS state_display,
        po.date_order::date AS order_date,
        po.date_planned::date AS planned_date,
        po.amount_total AS po_total,
        po.amount_untaxed AS po_subtotal,
        po.amount_tax AS po_tax,
        rp.name AS supplier_name,
        rp.vat AS supplier_nit,
        CASE 
            WHEN po.currency_id = 163 THEN 'GTQ'
            WHEN po.currency_id = 2 THEN 'USD'
            ELSE rc.name
        END AS currency,
        CASE 
            WHEN po.currency_id = 2 THEN po.amount_total * 7.7
            ELSE po.amount_total
        END AS amount_gtq,
        pot.task_id,
        pot.task_name,
        pot.start_date AS task_start_date,
        pot.end_date AS task_end_date,
        pot.project_name,
        pot.project_id,
        COALESCE(pot.start_date, po.date_planned, po.date_order) AS payment_date,
        CASE 
            WHEN po.invoice_status = 'no' THEN po.amount_total
            WHEN po.invoice_status = 'to invoice' THEN po.amount_total
            ELSE COALESCE(inv.unpaid_amount, 0)
        END AS amount_pending,
        po.internal_notes AS notes,
        po.invoice_status,
        po.receipt_status
        
    FROM purchase_order po
    JOIN res_partner rp ON po.partner_id = rp.id
    LEFT JOIN res_currency rc ON po.currency_id = rc.id
    LEFT JOIN po_tasks pot ON po.id = pot.po_id
    LEFT JOIN invoice_data inv ON po.id = inv.order_id
    
    WHERE po.state IN ('draft', 'sent', 'to approve', 'purchase')
    
    ORDER BY 
        COALESCE(pot.start_date, po.date_planned, po.date_order),
        po.name
    """
    
    print("Fetching cash flow data...")
    conn_str = f"postgresql://{DATABASE_CONFIG['user']}:{DATABASE_CONFIG['password']}@{DATABASE_CONFIG['host']}:{DATABASE_CONFIG['port']}/{DATABASE_CONFIG['dbname']}"
    engine = create_engine(conn_str)
    
    try:
        df = pd.read_sql_query(query, engine)
        print(f"Fetched {len(df)} purchase order records")
        df['project_name'] = df['project_name'].fillna('Sin Proyecto')
        df['task_name'] = df['task_name'].fillna('Sin Tarea')
        return df
    finally:
        engine.dispose()

def calculate_weekly_cash_flow(df, weeks_ahead=12):
    """Calculate weekly cash flow projections"""
    
    today = datetime.date.today()
    days_since_monday = today.weekday()
    current_week_start = today - datetime.timedelta(days=days_since_monday)
    
    weekly_data = {}
    
    for week_num in range(weeks_ahead + 1):
        week_start = current_week_start + datetime.timedelta(weeks=week_num)
        week_end = week_start + datetime.timedelta(days=6)
        week_key = week_start.strftime('%Y-%m-%d')
        
        weekly_data[week_key] = {
            'week_start': week_start,
            'week_end': week_end,
            'week_label': f"Semana {week_start.strftime('%d/%m')} - {week_end.strftime('%d/%m')}",
            'payments': [],
            'total_gtq': 0,
            'total_usd': 0,
            'count': 0,
            'by_status': {'draft': 0, 'sent': 0, 'purchase': 0},
            'by_project': {}
        }
    
    for _, row in df.iterrows():
        payment_date = pd.to_datetime(row['payment_date'])
        if pd.isna(payment_date):
            payment_date = today
        else:
            payment_date = payment_date.date()
        
        days_diff = (payment_date - current_week_start).days
        week_num = days_diff // 7
        
        if week_num < 0 or week_num > weeks_ahead:
            continue
        
        week_start = current_week_start + datetime.timedelta(weeks=week_num)
        week_key = week_start.strftime('%Y-%m-%d')
        
        payment_info = {
            'po_number': row['po_number'],
            'supplier': row['supplier_name'],
            'amount': row['amount_pending'],
            'amount_gtq': row['amount_gtq'],
            'currency': row['currency'],
            'status': row['po_state'],
            'project': row['project_name'],
            'task': row['task_name'],
            'payment_date': payment_date
        }
        
        weekly_data[week_key]['payments'].append(payment_info)
        weekly_data[week_key]['count'] += 1
        
        if row['currency'] == 'GTQ':
            weekly_data[week_key]['total_gtq'] += row['amount_pending']
        else:
            weekly_data[week_key]['total_usd'] += row['amount_pending']
            weekly_data[week_key]['total_gtq'] += row['amount_gtq']
        
        if row['po_state'] in weekly_data[week_key]['by_status']:
            weekly_data[week_key]['by_status'][row['po_state']] += row['amount_gtq']
        
        project = row['project_name']
        if project not in weekly_data[week_key]['by_project']:
            weekly_data[week_key]['by_project'][project] = 0
        weekly_data[week_key]['by_project'][project] += row['amount_gtq']
    
    return weekly_data


def clean_cell_value(value):
    """Clean a value for Excel compatibility - removes illegal XML/Excel characters"""
    if value is None:
        return ''
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value
    # Convert to string and strip illegal XML characters
    text = str(value)
    # Remove characters illegal in Excel XML (control chars except tab, newline, carriage return)
    cleaned = ''.join(
        char for char in text
        if ord(char) >= 32 and ord(char) != 127
        or char in ('\n', '\r', '\t')
    )
    # Remove Unicode surrogates and other problematic chars
    cleaned = cleaned.encode('utf-8', errors='replace').decode('utf-8')
    return cleaned


def create_cash_flow_excel(weekly_data, df, output_path):
    """Create comprehensive cash flow Excel report"""
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Color scheme
    header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    week_fill = PatternFill(start_color="f3f4f6", end_color="f3f4f6", fill_type="solid")
    total_fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
    
    # ── 1. WEEKLY SUMMARY TAB ──
    ws_summary = wb.create_sheet("Resumen Semanal")
    
    summary_headers = ['Semana', 'Fecha Inicio', 'Fecha Fin', '# Ordenes', 
                      'Total GTQ', 'Total USD', 'Borrador', 'Enviado', 'Confirmado']
    
    for c_idx, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=c_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    row_num = 2
    total_amounts = {'gtq': 0, 'usd': 0, 'count': 0}
    
    for week_key in sorted(weekly_data.keys()):
        week = weekly_data[week_key]
        
        ws_summary.cell(row=row_num, column=1, value=clean_cell_value(week['week_label']))
        ws_summary.cell(row=row_num, column=2, value=week['week_start'])
        ws_summary.cell(row=row_num, column=3, value=week['week_end'])
        ws_summary.cell(row=row_num, column=4, value=week['count'])
        ws_summary.cell(row=row_num, column=5, value=week['total_gtq'])
        ws_summary.cell(row=row_num, column=6, value=week['total_usd'])
        ws_summary.cell(row=row_num, column=7, value=week['by_status'].get('draft', 0))
        ws_summary.cell(row=row_num, column=8, value=week['by_status'].get('sent', 0))
        ws_summary.cell(row=row_num, column=9, value=week['by_status'].get('purchase', 0))
        
        for c in range(1, 10):
            if row_num % 2 == 0:
                ws_summary.cell(row=row_num, column=c).fill = week_fill
        
        for c in [5, 6, 7, 8, 9]:
            ws_summary.cell(row=row_num, column=c).number_format = '#,##0.00'
        
        for c in [2, 3]:
            ws_summary.cell(row=row_num, column=c).number_format = 'dd/mm/yyyy'
        
        total_amounts['gtq'] += week['total_gtq']
        total_amounts['usd'] += week['total_usd']
        total_amounts['count'] += week['count']
        
        row_num += 1
    
    # Totals row
    ws_summary.cell(row=row_num, column=1, value="TOTAL").font = Font(bold=True)
    ws_summary.cell(row=row_num, column=4, value=total_amounts['count'])
    ws_summary.cell(row=row_num, column=5, value=total_amounts['gtq'])
    ws_summary.cell(row=row_num, column=6, value=total_amounts['usd'])
    
    for c in range(1, 10):
        ws_summary.cell(row=row_num, column=c).fill = total_fill
        ws_summary.cell(row=row_num, column=c).font = Font(bold=True)
    
    # ── 2. DETAILED PAYMENTS TAB ──
    ws_detail = wb.create_sheet("Detalle de Pagos")
    
    detail_headers = ['Semana', 'PO Numero', 'Proveedor', 'Proyecto', 'Tarea', 
                      'Estado', 'Moneda', 'Monto', 'Monto GTQ', 'Fecha Pago']
    
    for c_idx, header in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=1, column=c_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    row_num = 2
    for week_key in sorted(weekly_data.keys()):
        week = weekly_data[week_key]
        for payment in week['payments']:
            ws_detail.cell(row=row_num, column=1, value=clean_cell_value(week['week_label']))
            ws_detail.cell(row=row_num, column=2, value=clean_cell_value(payment['po_number']))
            ws_detail.cell(row=row_num, column=3, value=clean_cell_value(payment['supplier']))
            ws_detail.cell(row=row_num, column=4, value=clean_cell_value(payment['project']))
            ws_detail.cell(row=row_num, column=5, value=clean_cell_value(payment['task']))
            ws_detail.cell(row=row_num, column=6, value=clean_cell_value(payment['status']))
            ws_detail.cell(row=row_num, column=7, value=clean_cell_value(payment['currency']))
            ws_detail.cell(row=row_num, column=8, value=payment['amount'])
            ws_detail.cell(row=row_num, column=9, value=payment['amount_gtq'])
            ws_detail.cell(row=row_num, column=10, value=clean_cell_value(payment['payment_date']))
            
            ws_detail.cell(row=row_num, column=8).number_format = '#,##0.00'
            ws_detail.cell(row=row_num, column=9).number_format = '#,##0.00'
            ws_detail.cell(row=row_num, column=10).number_format = 'dd/mm/yyyy'
            
            row_num += 1
    
    # ── 3. PROJECT SUMMARY TAB ──
    ws_project = wb.create_sheet("Por Proyecto")
    
    project_totals = {}
    for week_key, week in weekly_data.items():
        for project, amount in week['by_project'].items():
            if project not in project_totals:
                project_totals[project] = {'total': 0, 'weeks': {}}
            project_totals[project]['total'] += amount
            project_totals[project]['weeks'][week_key] = amount
    
    # Headers
    ws_project.cell(row=1, column=1, value="Proyecto").fill = header_fill
    ws_project.cell(row=1, column=1).font = header_font
    
    col_num = 2
    week_columns = {}
    for week_key in sorted(weekly_data.keys()):
        week = weekly_data[week_key]
        cell = ws_project.cell(row=1, column=col_num, value=clean_cell_value(week['week_label']))
        cell.fill = header_fill
        cell.font = header_font
        week_columns[week_key] = col_num
        col_num += 1
    
    ws_project.cell(row=1, column=col_num, value="Total").fill = header_fill
    ws_project.cell(row=1, column=col_num).font = header_font
    
    # Project data
    row_num = 2
    for project in sorted(project_totals.keys()):
        ws_project.cell(row=row_num, column=1, value=clean_cell_value(project))
        
        for week_key, col in week_columns.items():
            amount = project_totals[project]['weeks'].get(week_key, 0)
            cell = ws_project.cell(row=row_num, column=col, value=amount)
            cell.number_format = '#,##0.00'
        
        cell = ws_project.cell(row=row_num, column=col_num, value=project_totals[project]['total'])
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True)
        
        row_num += 1
    
    # Totals row
    ws_project.cell(row=row_num, column=1, value="TOTAL").font = Font(bold=True)
    for week_key, col in week_columns.items():
        total = sum(p['weeks'].get(week_key, 0) for p in project_totals.values())
        cell = ws_project.cell(row=row_num, column=col, value=total)
        cell.number_format = '#,##0.00'
        cell.fill = total_fill
        cell.font = Font(bold=True)
    
    grand_total = sum(p['total'] for p in project_totals.values())
    cell = ws_project.cell(row=row_num, column=col_num, value=grand_total)
    cell.number_format = '#,##0.00'
    cell.fill = total_fill
    cell.font = Font(bold=True)
    
    # Auto-adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column[:100]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(output_path)
    print(f"\nCash flow report saved to: {output_path}")


def create_json_summary(weekly_data, output_path):
    """Create JSON summary for the web viewer"""
    
    summary = {
        'generated_at': datetime.datetime.now().isoformat(),
        'weeks': [],
        'totals': {
            'total_gtq': 0,
            'total_usd': 0,
            'total_orders': 0
        }
    }
    
    for week_key in sorted(weekly_data.keys()):
        week = weekly_data[week_key]
        week_summary = {
            'week_start': week['week_start'].isoformat(),
            'week_end': week['week_end'].isoformat(),
            'label': week['week_label'],
            'total_gtq': week['total_gtq'],
            'total_usd': week['total_usd'],
            'count': week['count'],
            'by_status': week['by_status'],
            'top_projects': sorted(
                week['by_project'].items(), 
                key=lambda x: x[1], 
                reverse=True
            )[:5]
        }
        summary['weeks'].append(week_summary)
        
        summary['totals']['total_gtq'] += week['total_gtq']
        summary['totals']['total_usd'] += week['total_usd']
        summary['totals']['total_orders'] += week['count']
    
    with open(output_path, 'w') as f:
        json.dump(summary, f, indent=2, default=str)
    
    print(f"JSON summary saved to: {output_path}")


def main():
    """Main execution"""
    print("="*60)
    print("CASH FLOW PROJECTION REPORT")
    print(f"Generated at: {datetime.datetime.now()}")
    print("="*60)
    
    try:
        date_folder = datetime.datetime.now().strftime('%Y-%m-%d')
        output_dir = os.path.join(BASE_DIR, date_folder)
        os.makedirs(output_dir, exist_ok=True)
        
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_filename = f'Cash_Flow_Report_{timestamp}.xlsx'
        json_filename = f'cash_flow_summary_{timestamp}.json'
        
        excel_path = os.path.join(output_dir, excel_filename)
        json_path = os.path.join(output_dir, json_filename)
        
        print("\nFetching purchase order and project data...")
        df = get_cash_flow_data()
        
        if df.empty:
            print("No open purchase orders found.")
            return
        
        print("\nCalculating weekly cash flow projections...")
        weekly_data = calculate_weekly_cash_flow(df, weeks_ahead=12)
        
        print("\n" + "="*60)
        print("CASH FLOW SUMMARY")
        print("="*60)
        
        total_gtq = sum(w['total_gtq'] for w in weekly_data.values())
        total_usd = sum(w['total_usd'] for w in weekly_data.values())
        total_orders = sum(w['count'] for w in weekly_data.values())
        
        print(f"\nTotal Orders: {total_orders}")
        print(f"Total Amount GTQ: Q{total_gtq:,.2f}")
        print(f"Total Amount USD: ${total_usd:,.2f}")
        
        print("\nNext 4 Weeks:")
        for i, week_key in enumerate(sorted(weekly_data.keys())[:4]):
            week = weekly_data[week_key]
            print(f"  {week['week_label']}: {week['count']} orders, Q{week['total_gtq']:,.2f}")
        
        print("\nGenerating Excel report...")
        create_cash_flow_excel(weekly_data, df, excel_path)
        
        print("Generating JSON summary...")
        create_json_summary(weekly_data, json_path)
        
        print("\nCASH FLOW REPORT COMPLETED!")
        print(f"\nFiles generated:")
        print(f"  - Excel Report: {excel_filename}")
        print(f"  - JSON Summary: {json_filename}")
        print("\nThe Cash Flow Viewer can now display this data.")
        
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()