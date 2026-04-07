#!/usr/bin/env python3
"""
Daily Attendance Report for Torre Los Beaterios
Real-time tracking of who's on site, attendance, and security control
Fixed version with proper timezone handling
"""

import pandas as pd
from sqlalchemy import create_engine, text
from datetime import datetime, timedelta
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import json
import sys

# Import database configuration from config.py
try:
    from config import DATABASE_CONFIG, BASE_DIR, WEB_DIR
except ImportError:
    DATABASE_CONFIG = {
        "dbname": "master",
        "user": "rudy",
        "password": "Kaibil10!z",
        "host": "3.18.71.107",
        "port": "5432"
    }
    BASE_DIR = "/opt/torelo-kpi/reports"
    WEB_DIR = "/opt/torelo-kpi/web"

# Guatemala timezone
GUATEMALA_TZ = pytz.timezone('America/Guatemala')
UTC_TZ = pytz.UTC

# Constants
DAILY_REGULAR_HOURS = 8
WEEKLY_REGULAR_HOURS = 44

# Departments to track (construcción y seguridad)
TRACKED_DEPARTMENTS = [
    'Obra',
    'Seguridad Industrial', 
    'Operativo',
    'Bodega',
    'Administración'
]

def safe_tz_convert(dt_series, target_tz):
    """Safely convert datetime series to target timezone"""
    if dt_series.empty or dt_series.isna().all():
        return dt_series
    
    # Make a copy to avoid modifying original
    result = dt_series.copy()
    
    # Check if tz-naive and localize if needed
    if result.dt.tz is None:
        result = result.dt.tz_localize('UTC')
    
    # Convert to target timezone
    return result.dt.tz_convert(target_tz)

def connect_to_db():
    """Create database connection"""
    db_params = DATABASE_CONFIG
    conn_str = f"postgresql://{db_params['user']}:{db_params['password']}@{db_params['host']}:{db_params['port']}/{db_params['dbname']}"
    return create_engine(conn_str)

def fetch_site_employees(engine):
    """Fetch ONLY employees who work on site (not all 80+)"""
    query = """
    SELECT DISTINCT
        e.id as employee_id,
        e.name as employee_name,
        d.name as department_name,
        e.job_title,
        e.mobile_phone as phone,
        e.pin as access_pin,
        e.barcode as employee_code
    FROM hr_employee e
    LEFT JOIN hr_department d ON e.department_id = d.id
    WHERE e.active = TRUE
        AND d.name IN :departments
    ORDER BY d.name, e.name
    """
    
    df = pd.read_sql_query(
        text(query), 
        engine,
        params={"departments": tuple(TRACKED_DEPARTMENTS)}
    )
    print(f"📊 Found {len(df)} employees in construction site departments")
    return df

def fetch_currently_on_site(engine, date):
    """CRITICAL: Find who is currently INSIDE the construction site"""
    query = """
    WITH last_actions AS (
        SELECT DISTINCT ON (employee_id)
            employee_id,
            check_in,
            check_out,
            CASE 
                WHEN check_out IS NULL THEN 'INSIDE'
                WHEN check_out > check_in THEN 'OUTSIDE'
                ELSE 'UNKNOWN'
            END as current_status
        FROM hr_attendance
        WHERE DATE(check_in AT TIME ZONE 'UTC' AT TIME ZONE 'America/Guatemala') = :target_date
        ORDER BY employee_id, check_in DESC
    )
    SELECT 
        la.employee_id,
        e.name as employee_name,
        d.name as department_name,
        la.check_in as entry_time_utc,
        la.check_out as exit_time_utc,
        la.current_status,
        CASE 
            WHEN la.check_out IS NULL THEN 
                EXTRACT(EPOCH FROM (NOW() - la.check_in))/3600
            ELSE 0
        END as hours_on_site
    FROM last_actions la
    JOIN hr_employee e ON e.id = la.employee_id
    LEFT JOIN hr_department d ON e.department_id = d.id
    WHERE la.current_status = 'INSIDE'
        AND d.name IN :departments
    ORDER BY la.check_in
    """
    
    df = pd.read_sql_query(
        text(query),
        engine,
        params={
            "target_date": date,
            "departments": tuple(TRACKED_DEPARTMENTS)
        }
    )
    
    # Convert UTC times to Guatemala time with safe conversion
    if not df.empty:
        df['entry_time_utc'] = pd.to_datetime(df['entry_time_utc'])
        df['entry_time'] = safe_tz_convert(df['entry_time_utc'], GUATEMALA_TZ)
        
        if 'exit_time_utc' in df.columns:
            df['exit_time_utc'] = pd.to_datetime(df['exit_time_utc'])
            mask = df['exit_time_utc'].notna()
            if mask.any():
                df.loc[mask, 'exit_time'] = safe_tz_convert(df.loc[mask, 'exit_time_utc'], GUATEMALA_TZ)
    
    print(f"⚠️ {len(df)} people currently INSIDE the construction site")
    return df

def fetch_period_attendance(engine, start_date, end_date):
    """Fetch attendance for a date range with real names"""
    query = """
    SELECT 
        a.employee_id,
        e.name as employee_name,
        d.name as department_name,
        DATE(a.check_in AT TIME ZONE 'UTC' AT TIME ZONE 'America/Guatemala') as work_date,
        MIN(a.check_in) as first_entry_utc,
        MAX(a.check_out) as last_exit_utc,
        COUNT(DISTINCT DATE(a.check_in AT TIME ZONE 'UTC' AT TIME ZONE 'America/Guatemala')) as days_present,
        SUM(
            CASE 
                WHEN a.check_out IS NOT NULL THEN 
                    EXTRACT(EPOCH FROM (a.check_out - a.check_in))/3600
                ELSE 0
            END
        ) as total_hours,
        ARRAY_AGG(
            CASE 
                WHEN a.check_out IS NULL THEN 'NO_EXIT_RECORDED'
                ELSE NULL
            END
        ) as missing_exits
    FROM hr_attendance a
    JOIN hr_employee e ON e.id = a.employee_id
    LEFT JOIN hr_department d ON e.department_id = d.id
    WHERE DATE(a.check_in AT TIME ZONE 'UTC' AT TIME ZONE 'America/Guatemala') >= :start_date 
        AND DATE(a.check_in AT TIME ZONE 'UTC' AT TIME ZONE 'America/Guatemala') <= :end_date
        AND d.name IN :departments
    GROUP BY a.employee_id, e.name, d.name, DATE(a.check_in AT TIME ZONE 'UTC' AT TIME ZONE 'America/Guatemala')
    ORDER BY e.name, work_date
    """
    
    df = pd.read_sql_query(
        text(query),
        engine,
        params={
            "start_date": start_date,
            "end_date": end_date,
            "departments": tuple(TRACKED_DEPARTMENTS)
        }
    )
    
    # Convert UTC times to Guatemala time with safe conversion
    if not df.empty:
        if 'first_entry_utc' in df.columns:
            df['first_entry_utc'] = pd.to_datetime(df['first_entry_utc'])
            df['first_entry'] = safe_tz_convert(df['first_entry_utc'], GUATEMALA_TZ)
        
        if 'last_exit_utc' in df.columns:
            df['last_exit_utc'] = pd.to_datetime(df['last_exit_utc'])
            mask = df['last_exit_utc'].notna()
            if mask.any():
                df.loc[mask, 'last_exit'] = safe_tz_convert(df.loc[mask, 'last_exit_utc'], GUATEMALA_TZ)
    
    return df

def fetch_daily_movements(engine, date):
    """Get all IN/OUT movements for the day with fixed timezone handling"""
    query = """
    SELECT 
        a.employee_id,
        e.name as employee_name,
        d.name as department_name,
        a.check_in as entry_time_utc,
        a.check_out as exit_time_utc,
        CASE 
            WHEN a.check_out IS NULL THEN 'STILL_INSIDE'
            ELSE 'COMPLETED'
        END as status
    FROM hr_attendance a
    JOIN hr_employee e ON e.id = a.employee_id
    LEFT JOIN hr_department d ON e.department_id = d.id
    WHERE DATE(a.check_in AT TIME ZONE 'UTC' AT TIME ZONE 'America/Guatemala') = :target_date
        AND d.name IN :departments
    ORDER BY a.check_in DESC
    """
    
    df = pd.read_sql_query(
        text(query),
        engine,
        params={
            "target_date": date,
            "departments": tuple(TRACKED_DEPARTMENTS)
        }
    )
    
    # Convert UTC times to Guatemala time with proper timezone handling
    if not df.empty:
        # Handle entry times
        df['entry_time_utc'] = pd.to_datetime(df['entry_time_utc'])
        df['entry_time'] = safe_tz_convert(df['entry_time_utc'], GUATEMALA_TZ)
        
        # Handle exit times (may contain NaN/None values)
        df['exit_time_utc'] = pd.to_datetime(df['exit_time_utc'])
        df['exit_time'] = pd.NaT  # Initialize with NaT
        
        # Only process non-null exit times
        mask = df['exit_time_utc'].notna()
        if mask.any():
            df.loc[mask, 'exit_time'] = safe_tz_convert(df.loc[mask, 'exit_time_utc'], GUATEMALA_TZ)
    
    return df

#!/usr/bin/env python3
"""
Add this function to your existing daily_attendance_site_security.py script
to generate comprehensive attendance data in JSON format
"""

def generate_comprehensive_attendance_json(engine, output_dir):
    """
    Generate a comprehensive JSON file with all attendance data
    This allows the HTML dashboard to load everything once and filter client-side
    """
    import json
    from datetime import datetime, timedelta
    import pandas as pd
    
    # Get the last 30 days of data
    end_date = datetime.now(GUATEMALA_TZ).date()
    start_date = end_date - timedelta(days=30)
    
    print("📊 Generating comprehensive attendance JSON...")
    
    # Fetch all attendance records for the period
    query = """
    SELECT 
        a.id as record_id,
        a.employee_id,
        e.name as employee_name,
        e.job_title,
        e.mobile_phone as phone,
        d.name as department_name,
        a.check_in as check_in_utc,
        a.check_out as check_out_utc,
        DATE(a.check_in AT TIME ZONE 'UTC' AT TIME ZONE 'America/Guatemala') as work_date
    FROM hr_attendance a
    JOIN hr_employee e ON e.id = a.employee_id
    LEFT JOIN hr_department d ON e.department_id = d.id
    WHERE DATE(a.check_in AT TIME ZONE 'UTC' AT TIME ZONE 'America/Guatemala') >= :start_date 
        AND DATE(a.check_in AT TIME ZONE 'UTC' AT TIME ZONE 'America/Guatemala') <= :end_date
        AND d.name IN :departments
    ORDER BY work_date DESC, a.check_in
    """
    
    df = pd.read_sql_query(
        text(query),
        engine,
        params={
            "start_date": start_date,
            "end_date": end_date,
            "departments": tuple(TRACKED_DEPARTMENTS)
        }
    )
    
    # Convert times to Guatemala timezone
    if not df.empty:
        df['check_in_utc'] = pd.to_datetime(df['check_in_utc'])
        df['check_in'] = safe_tz_convert(df['check_in_utc'], GUATEMALA_TZ)
        
        df['check_out_utc'] = pd.to_datetime(df['check_out_utc'])
        mask = df['check_out_utc'].notna()
        df['check_out'] = pd.NaT
        if mask.any():
            df.loc[mask, 'check_out'] = safe_tz_convert(df.loc[mask, 'check_out_utc'], GUATEMALA_TZ)
    
    # Group by date and employee to organize records
    attendance_by_date = {}
    
    for date in df['work_date'].unique():
        date_str = str(date)
        date_records = df[df['work_date'] == date]
        
        # Group by employee for this date
        employees_data = []
        for emp_id in date_records['employee_id'].unique():
            emp_records = date_records[date_records['employee_id'] == emp_id].sort_values('check_in')
            
            # Get all clock ins/outs for this employee on this date
            clock_records = []
            total_hours = 0
            lunch_duration = 0
            
            for _, record in emp_records.iterrows():
                check_in_time = record['check_in'].strftime('%H:%M:%S') if pd.notna(record['check_in']) else None
                check_out_time = record['check_out'].strftime('%H:%M:%S') if pd.notna(record['check_out']) else None
                
                clock_records.append({
                    'type': 'entry',
                    'time': check_in_time
                })
                
                if check_out_time:
                    clock_records.append({
                        'type': 'exit',
                        'time': check_out_time
                    })
                    
                    # Calculate hours for this segment
                    duration = (record['check_out'] - record['check_in']).total_seconds() / 3600
                    total_hours += duration
            
            # Detect lunch breaks (gap between clock out and next clock in)
            if len(emp_records) > 1:
                for i in range(len(emp_records) - 1):
                    current_out = emp_records.iloc[i]['check_out']
                    next_in = emp_records.iloc[i + 1]['check_in']
                    if pd.notna(current_out) and pd.notna(next_in):
                        lunch_duration += (next_in - current_out).total_seconds() / 3600
            
            # Determine current status
            last_record = emp_records.iloc[-1]
            is_currently_inside = pd.isna(last_record['check_out'])
            
            # Calculate effective hours (total minus lunch)
            effective_hours = total_hours
            
            employees_data.append({
                'employee_id': int(emp_id),
                'employee_name': emp_records.iloc[0]['employee_name'],
                'department': emp_records.iloc[0]['department_name'],
                'job_title': emp_records.iloc[0]['job_title'],
                'phone': emp_records.iloc[0]['phone'],
                'clock_records': clock_records,
                'total_hours': round(total_hours, 2),
                'lunch_duration': round(lunch_duration, 2),
                'effective_hours': round(effective_hours, 2),
                'is_currently_inside': is_currently_inside,
                'first_entry': emp_records.iloc[0]['check_in'].strftime('%H:%M:%S') if pd.notna(emp_records.iloc[0]['check_in']) else None,
                'last_exit': emp_records.iloc[-1]['check_out'].strftime('%H:%M:%S') if pd.notna(emp_records.iloc[-1]['check_out']) else None
            })
        
        attendance_by_date[date_str] = {
            'date': date_str,
            'total_employees': len(employees_data),
            'currently_inside': sum(1 for e in employees_data if e['is_currently_inside']),
            'employees': employees_data
        }
    
    # Get all employees list
    all_employees_query = """
    SELECT DISTINCT
        e.id as employee_id,
        e.name as employee_name,
        d.name as department_name,
        e.job_title,
        e.mobile_phone as phone
    FROM hr_employee e
    LEFT JOIN hr_department d ON e.department_id = d.id
    WHERE e.active = TRUE
        AND d.name IN :departments
    ORDER BY e.name
    """
    
    employees_df = pd.read_sql_query(
        text(all_employees_query),
        engine,
        params={"departments": tuple(TRACKED_DEPARTMENTS)}
    )
    
    all_employees = []
    for _, emp in employees_df.iterrows():
        all_employees.append({
            'id': int(emp['employee_id']),
            'name': emp['employee_name'],
            'department': emp['department_name'],
            'job_title': emp['job_title'],
            'phone': emp['phone']
        })
    
    # Create the comprehensive data structure
    comprehensive_data = {
        'generated_at': datetime.now(GUATEMALA_TZ).isoformat(),
        'timezone': 'America/Guatemala',
        'data_range': {
            'start': start_date.isoformat(),
            'end': end_date.isoformat()
        },
        'all_employees': all_employees,
        'attendance_by_date': attendance_by_date,
        'summary': {
            'total_employees': len(all_employees),
            'total_days': len(attendance_by_date),
            'departments': list(TRACKED_DEPARTMENTS)
        }
    }
    
    # Save to file
    output_file = os.path.join(output_dir, 'comprehensive_attendance.json')
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(comprehensive_data, f, indent=2, ensure_ascii=False, default=str)
    
    print(f"✅ Comprehensive attendance data saved to: {output_file}")
    
    # Also copy to web directory for easy access
    web_file = os.path.join(WEB_DIR, 'latest', 'comprehensive_attendance.json')
    if os.path.exists(os.path.dirname(web_file)):
        import shutil
        shutil.copy2(output_file, web_file)
        print(f"📋 Copied to web directory: {web_file}")
    
    return comprehensive_data

# Add this to your main() function in the existing script:
def main(target_date=None):
    # ... existing code ...
    
    # After generating the regular reports, add:
    print("\n📊 Generating comprehensive attendance JSON...")
    comprehensive_data = generate_comprehensive_attendance_json(engine, today_reports)
    
    # ... rest of existing code ...

def create_security_report(employees_df, currently_inside, daily_movements, period_data, 
                          report_date, output_file):
    """Create comprehensive security and attendance report"""
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")
    danger_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    warning_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    safe_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: CURRENTLY ON SITE (CRITICAL)
    ws_onsite = wb.active
    ws_onsite.title = "EN SITIO AHORA"
    
    # Header
    ws_onsite.merge_cells('A1:H1')
    ws_onsite['A1'] = f"⚠️ PERSONAL ACTUALMENTE EN OBRA - {datetime.now(GUATEMALA_TZ).strftime('%H:%M:%S')}"
    ws_onsite['A1'].font = Font(bold=True, size=16, color="FF0000")
    ws_onsite['A1'].fill = warning_fill
    ws_onsite['A1'].alignment = Alignment(horizontal='center')
    
    ws_onsite.merge_cells('A2:H2')
    ws_onsite['A2'] = f"Torre Los Beaterios - {report_date.strftime('%d/%m/%Y')}"
    ws_onsite['A2'].font = Font(bold=True, size=12)
    ws_onsite['A2'].alignment = Alignment(horizontal='center')
    
    # Summary
    ws_onsite['A4'] = f"TOTAL EN SITIO: {len(currently_inside)}"
    ws_onsite['A4'].font = Font(bold=True, size=14, color="FF0000")
    
    # Emergency contact info
    ws_onsite['A5'] = "EMERGENCIA: 911 | Bomberos: 122 | Proyecto: +502 XXXX-XXXX"
    ws_onsite['A5'].font = Font(italic=True, color="FF0000")
    
    # Table headers
    row = 7
    headers = ["No.", "NOMBRE", "DEPARTAMENTO", "HORA ENTRADA", "TIEMPO EN SITIO", "TELÉFONO", "ESTADO", "ALERTA"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_onsite.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row = 8
    for idx, person in currently_inside.iterrows():
        ws_onsite.cell(row=row, column=1, value=idx + 1).border = border
        
        # Name in BOLD
        name_cell = ws_onsite.cell(row=row, column=2, value=person['employee_name'])
        name_cell.font = Font(bold=True)
        name_cell.border = border
        
        ws_onsite.cell(row=row, column=3, value=person['department_name']).border = border
        
        # Entry time
        if pd.notna(person.get('entry_time')):
            ws_onsite.cell(row=row, column=4, value=person['entry_time'].strftime('%H:%M')).border = border
        
        # Hours on site
        hours_on_site = person.get('hours_on_site', 0)
        hours_cell = ws_onsite.cell(row=row, column=5, value=f"{hours_on_site:.1f} horas")
        hours_cell.border = border
        
        # Alert based on time on site
        alert_cell = ws_onsite.cell(row=row, column=8)
        if hours_on_site > 12:
            hours_cell.fill = danger_fill
            alert_cell.value = "⚠️ EXCESO"
            alert_cell.fill = danger_fill
            alert_cell.font = Font(bold=True, color="FFFFFF")
        elif hours_on_site > 10:
            hours_cell.fill = warning_fill
            alert_cell.value = "PROLONGADO"
            alert_cell.fill = warning_fill
        else:
            alert_cell.value = "NORMAL"
        alert_cell.border = border
        
        # Phone (from employees_df)
        emp_info = employees_df[employees_df['employee_id'] == person['employee_id']]
        if not emp_info.empty:
            ws_onsite.cell(row=row, column=6, value=emp_info.iloc[0]['phone'] or 'N/A').border = border
        
        # Status
        status_cell = ws_onsite.cell(row=row, column=7, value="EN SITIO")
        status_cell.fill = warning_fill
        status_cell.border = border
        
        row += 1
    
    # Sheet 2: Daily Movements (All IN/OUT)
    ws_moves = wb.create_sheet("Movimientos del Día")
    
    ws_moves.merge_cells('A1:G1')
    ws_moves['A1'] = f"REGISTRO DE ENTRADAS Y SALIDAS - {report_date.strftime('%d/%m/%Y')}"
    ws_moves['A1'].font = Font(bold=True, size=14)
    ws_moves['A1'].alignment = Alignment(horizontal='center')
    
    # Stats
    total_entries = len(daily_movements)
    still_inside = len(daily_movements[daily_movements['status'] == 'STILL_INSIDE'])
    completed = len(daily_movements[daily_movements['status'] == 'COMPLETED'])
    
    ws_moves['A3'] = f"Total Entradas: {total_entries}"
    ws_moves['B3'] = f"Salidas Completas: {completed}"
    ws_moves['C3'] = f"AÚN ADENTRO: {still_inside}"
    ws_moves['C3'].font = Font(bold=True, color="FF0000")
    
    # Table
    row = 5
    headers = ["No.", "NOMBRE", "DEPTO", "ENTRADA", "SALIDA", "DURACIÓN", "ESTADO"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_moves.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
        cell.border = border
    
    row = 6
    for idx, move in daily_movements.iterrows():
        ws_moves.cell(row=row, column=1, value=idx + 1).border = border
        ws_moves.cell(row=row, column=2, value=move['employee_name']).border = border
        ws_moves.cell(row=row, column=3, value=move['department_name']).border = border
        
        # Entry
        if pd.notna(move.get('entry_time')):
            ws_moves.cell(row=row, column=4, value=move['entry_time'].strftime('%H:%M:%S')).border = border
        
        # Exit
        if pd.notna(move.get('exit_time')):
            ws_moves.cell(row=row, column=5, value=move['exit_time'].strftime('%H:%M:%S')).border = border
            # Duration
            duration = (move['exit_time'] - move['entry_time']).total_seconds() / 3600
            ws_moves.cell(row=row, column=6, value=f"{duration:.1f}h").border = border
        else:
            exit_cell = ws_moves.cell(row=row, column=5, value="NO SALIÓ")
            exit_cell.fill = danger_fill
            exit_cell.border = border
            ws_moves.cell(row=row, column=6, value="EN SITIO").border = border
        
        # Status
        status_cell = ws_moves.cell(row=row, column=7, value=move['status'])
        if move['status'] == 'STILL_INSIDE':
            status_cell.fill = warning_fill
            status_cell.font = Font(bold=True)
        else:
            status_cell.fill = safe_fill
        status_cell.border = border
        
        row += 1
    
    # Sheet 3: Period Summary
    ws_period = wb.create_sheet("Resumen Período")
    
    # Get week dates
    week_start = report_date - timedelta(days=report_date.weekday())
    week_end = week_start + timedelta(days=6)
    
    ws_period.merge_cells('A1:F1')
    ws_period['A1'] = f"RESUMEN SEMANAL: {week_start.strftime('%d/%m')} - {week_end.strftime('%d/%m/%Y')}"
    ws_period['A1'].font = Font(bold=True, size=14)
    ws_period['A1'].alignment = Alignment(horizontal='center')
    
    # Group by employee
    if not period_data.empty:
        employee_summary = period_data.groupby(['employee_id', 'employee_name', 'department_name']).agg({
            'days_present': 'sum',
            'total_hours': 'sum',
            'missing_exits': lambda x: sum(1 for exits in x if 'NO_EXIT_RECORDED' in str(exits))
        }).reset_index()
        
        row = 3
        headers = ["NOMBRE", "DEPARTAMENTO", "DÍAS", "HORAS TOTAL", "HORAS EXTRA", "SALIDAS FALTANTES"]
        
        for col, header in enumerate(headers, 1):
            cell = ws_period.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        row = 4
        for _, emp in employee_summary.iterrows():
            ws_period.cell(row=row, column=1, value=emp['employee_name']).border = border
            ws_period.cell(row=row, column=2, value=emp['department_name']).border = border
            ws_period.cell(row=row, column=3, value=emp['days_present']).border = border
            ws_period.cell(row=row, column=4, value=f"{emp['total_hours']:.1f}").border = border
            
            # Overtime
            overtime = max(0, emp['total_hours'] - WEEKLY_REGULAR_HOURS)
            ot_cell = ws_period.cell(row=row, column=5, value=f"{overtime:.1f}")
            if overtime > 0:
                ot_cell.fill = warning_fill
            ot_cell.border = border
            
            # Missing exits (security issue)
            if emp['missing_exits'] > 0:
                missing_cell = ws_period.cell(row=row, column=6, value=f"⚠️ {emp['missing_exits']}")
                missing_cell.fill = danger_fill
                missing_cell.font = Font(bold=True, color="FFFFFF")
            else:
                ws_period.cell(row=row, column=6, value="OK")
            ws_period.cell(row=row, column=6).border = border
            
            row += 1
    
    # Adjust column widths
    for ws in [ws_onsite, ws_moves, ws_period]:
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Save
    wb.save(output_file)
    
    return {
        'total_employees': len(employees_df),
        'currently_inside': len(currently_inside),
        'total_movements': len(daily_movements),
        'missing_exits': len(daily_movements[daily_movements['status'] == 'STILL_INSIDE'])
    }

def main(target_date=None):
    """Main function with improved error handling"""
    
    if target_date:
        if isinstance(target_date, str):
            report_date = datetime.strptime(target_date, '%Y-%m-%d').date()
        else:
            report_date = target_date
    else:
        report_date = datetime.now(GUATEMALA_TZ).date()
    
    print(f"\n{'='*60}")
    print(f"🚨 CONTROL DE SEGURIDAD Y ASISTENCIA - TORRE LOS BEATERIOS")
    print(f"{'='*60}")
    print(f"📅 Fecha: {report_date.strftime('%d/%m/%Y')}")
    print(f"⏰ Hora: {datetime.now(GUATEMALA_TZ).strftime('%H:%M:%S')}")
    print(f"🌐 Zona Horaria: Guatemala (UTC-6)")
    print(f"{'='*60}\n")
    
    engine = connect_to_db()
    
    try:
        # Get site employees only
        print("👥 Obteniendo empleados del sitio de construcción...")
        employees_df = fetch_site_employees(engine)
        
        # CRITICAL: Who's inside right now?
        print("\n🚨 VERIFICANDO QUIÉN ESTÁ EN EL SITIO AHORA...")
        currently_inside = fetch_currently_on_site(engine, report_date)
        
        if len(currently_inside) > 0:
            print("\n⚠️ PERSONAL ACTUALMENTE EN OBRA:")
            for _, person in currently_inside.iterrows():
                entry_time = person['entry_time'].strftime('%H:%M') if pd.notna(person.get('entry_time')) else 'N/A'
                hours = person.get('hours_on_site', 0)
                alert = " ⚠️ EXCESO!" if hours > 12 else " ⚠️" if hours > 10 else ""
                print(f"   - {person['employee_name']} ({person['department_name']}) - Entrada: {entry_time} - {hours:.1f}h{alert}")
        else:
            print("✅ No hay personal en obra en este momento")
        
        # Get all movements
        print("\n📊 Obteniendo movimientos del día...")
        daily_movements = fetch_daily_movements(engine, report_date)
        print(f"   Total movimientos: {len(daily_movements)}")
        
        # Get week data
        week_start = report_date - timedelta(days=report_date.weekday())
        week_end = report_date
        
        print(f"📅 Obteniendo datos del período...")
        period_data = fetch_period_attendance(engine, week_start, week_end)
        
        # Create output
        date_folder = report_date.strftime('%Y-%m-%d')
        today_reports = os.path.join(BASE_DIR, date_folder)
        os.makedirs(today_reports, exist_ok=True)
        
        output_file = os.path.join(
            today_reports,
            f"control_obra_{report_date.strftime('%Y%m%d')}_{datetime.now(GUATEMALA_TZ).strftime('%H%M')}.xlsx"
        )
        
        print("\n📝 Generando reporte de seguridad...")
        stats = create_security_report(
            employees_df, currently_inside, daily_movements, period_data,
            report_date, output_file
        )
        
        # Print summary
        print("\n" + "="*60)
        print("✅ REPORTE GENERADO EXITOSAMENTE")
        print("="*60)
        print(f"👥 Empleados del Sitio: {stats['total_employees']}")
        print(f"🚨 ACTUALMENTE EN OBRA: {stats['currently_inside']}")
        print(f"📊 Movimientos Hoy: {stats['total_movements']}")
        if stats['missing_exits'] > 0:
            print(f"⚠️ SIN REGISTRO DE SALIDA: {stats['missing_exits']}")
        print("="*60)
        print(f"\n📁 Archivo: {output_file}")
        
        # Save JSON for dashboard
        status_file = os.path.join(today_reports, "site_status.json")
        
        # Prepare data for JSON with proper serialization
        on_site_list = []
        if not currently_inside.empty:
            for _, person in currently_inside.iterrows():
                on_site_list.append({
                    'employee_id': int(person['employee_id']),
                    'employee_name': person['employee_name'],
                    'department_name': person['department_name'],
                    'entry_time': person['entry_time'].isoformat() if pd.notna(person.get('entry_time')) else None,
                    'hours_on_site': round(person.get('hours_on_site', 0), 1),
                    'alert_level': 'danger' if person.get('hours_on_site', 0) > 12 else 'warning' if person.get('hours_on_site', 0) > 10 else 'normal'
                })
        
        status = {
            "timestamp": datetime.now(GUATEMALA_TZ).isoformat(),
            "report_date": report_date.isoformat(),
            "currently_on_site": len(currently_inside),
            "on_site_list": on_site_list,
            "stats": stats,
            "file": os.path.basename(output_file),
            "emergency_numbers": {
                "bomberos": "122",
                "emergencia": "911",
                "policia": "110",
                "project_manager": "+502-XXXX-XXXX"
            }
        }
        
        with open(status_file, 'w') as f:
            json.dump(status, f, indent=2, default=str)
        
        # Copy to web
        latest_dir = os.path.join(WEB_DIR, 'latest')
        if os.path.exists(latest_dir):
            import shutil
            shutil.copy2(output_file, latest_dir)
            shutil.copy2(status_file, latest_dir)
            print(f"📋 Copiado a directorio web: {latest_dir}")
        
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        engine.dispose()


if __name__ == "__main__":
    if len(sys.argv) > 1:
        main(sys.argv[1])
    else:
        main()