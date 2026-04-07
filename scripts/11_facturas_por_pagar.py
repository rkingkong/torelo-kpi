#!/usr/bin/env python3
"""
Vendor Bills + PO Link Report
- Pulls account.move (vendor bills) + account.move.line
- Links to purchase orders when possible
- Shows what's due next 7 days, what's unpaid, and what's already paid
"""

import sys
sys.path.append('/opt/torelo-kpi/scripts')

from config import DATABASE_CONFIG, BASE_DIR
import pandas as pd
from sqlalchemy import create_engine
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import datetime
import os
import re
import html

# -----------------------------
# Helpers
# -----------------------------
def clean_html_text(text):
    if not text:
        return ""
    text = html.unescape(str(text))
    text = re.sub(r'<br\s*/?>', '\n', text, flags=re.IGNORECASE)
    text = re.sub(r'</p>', '\n', text, flags=re.IGNORECASE)
    text = re.sub(r'<[^>]+>', '', text)
    text = re.sub(r'\n\s*\n', '\n', text)
    text = re.sub(r'^\s+|\s+$', '', text)
    return text.strip()

def clean_excel_value(value):
    if value is None:
        return ""
    value = str(value)
    illegal_chars = [chr(i) for i in range(0, 32) if i not in [9, 10, 13]]
    for ch in illegal_chars:
        value = value.replace(ch, '')
    value = value.replace('\x00', '')
    value = value.replace('\r\n', '\n').replace('\r', '\n')
    lines = value.split('\n')
    lines = [' '.join(line.split()) for line in lines]
    return '\n'.join([l for l in lines if l]).strip()

def safe_sheet_title(title: str) -> str:
    """
    Excel forbids: : \ / ? * [ ]
    Max length: 31
    """
    title = re.sub(r'[:\\\\/\\?\\*\\[\\]]', '-', title)
    title = title.strip()
    return title[:31] if len(title) > 31 else title

def db_engine():
    conn_str = (
        f"postgresql://{DATABASE_CONFIG['user']}:{DATABASE_CONFIG['password']}"
        f"@{DATABASE_CONFIG['host']}:{DATABASE_CONFIG['port']}/{DATABASE_CONFIG['dbname']}"
    )
    return create_engine(conn_str)

# -----------------------------
# Core query
# -----------------------------
def get_vendor_bill_data(only_posted=True):
    """
    Returns a dataframe at invoice-level + line-level columns.
    We link POs via:
      1) account_move_purchase_order_rel
      2) aml.purchase_line_id -> purchase_order_line -> purchase_order
    """
    state_clause = "am.state = 'posted'" if only_posted else "am.state IN ('draft','posted')"

    query = f"""
    WITH po_from_rel AS (
        SELECT
            rel.account_move_id AS move_id,
            string_agg(DISTINCT po.name, ', ' ORDER BY po.name) AS po_numbers
        FROM account_move_purchase_order_rel rel
        JOIN purchase_order po ON po.id = rel.purchase_order_id
        GROUP BY rel.account_move_id
    ),
    po_from_lines AS (
        SELECT
            aml.move_id,
            string_agg(DISTINCT po.name, ', ' ORDER BY po.name) AS po_numbers
        FROM account_move_line aml
        JOIN purchase_order_line pol ON pol.id = aml.purchase_line_id
        JOIN purchase_order po ON po.id = pol.order_id
        GROUP BY aml.move_id
    ),
    po_links AS (
        SELECT
            m.id AS move_id,
            COALESCE(r.po_numbers, l.po_numbers) AS po_numbers
        FROM account_move m
        LEFT JOIN po_from_rel r ON r.move_id = m.id
        LEFT JOIN po_from_lines l ON l.move_id = m.id
    )
    SELECT
        -- Invoice (Vendor Bill) header
        am.id AS bill_id,
        am.name AS bill_number,
        am.move_type,
        am.state AS bill_state,
        am.payment_state,
        am.ref AS vendor_reference,
        am.invoice_origin,
        am.invoice_date,
        am.invoice_date_due,
        am.date AS accounting_date,
        am.amount_total_signed AS bill_total_signed,
        am.amount_residual_signed AS bill_residual_signed,

        -- Currency
        CASE
            WHEN am.currency_id = 163 THEN 'GTQ'
            WHEN am.currency_id = 2 THEN 'USD'
            ELSE rc.name
        END AS currency,

        -- Supplier
        rp.name AS supplier_name,
        rp.phone AS supplier_phone,
        rp.email AS supplier_email,

        -- Notes / narrative
        COALESCE(am.narration, '') AS bill_notes,

        -- PO link (if any)
        COALESCE(pl.po_numbers, '') AS po_numbers,

        -- Computed
        (am.amount_total_signed - am.amount_residual_signed) AS amount_paid_signed,
        COALESCE(am.invoice_date_due, am.invoice_date, am.date) AS effective_due_date,
        CASE
            WHEN (COALESCE(am.invoice_date_due, am.invoice_date, am.date) < CURRENT_DATE)
                 AND (am.amount_residual_signed > 0)
            THEN (CURRENT_DATE - COALESCE(am.invoice_date_due, am.invoice_date, am.date))
            ELSE 0
        END AS days_overdue,

        -- Lines
        aml.id AS line_id,
        aml.name AS line_label,
        aml.display_type,
        aml.quantity,
        aml.price_unit,
        aml.price_subtotal,
        aml.price_total,
        aml.ref AS line_ref,
        pt.name AS product_name,
        pp.default_code AS product_code

    FROM account_move am
    JOIN res_partner rp ON rp.id = am.partner_id
    LEFT JOIN res_currency rc ON rc.id = am.currency_id
    LEFT JOIN po_links pl ON pl.move_id = am.id

    LEFT JOIN account_move_line aml
        ON aml.move_id = am.id

    LEFT JOIN product_product pp ON pp.id = aml.product_id
    LEFT JOIN product_template pt ON pt.id = pp.product_tmpl_id

    WHERE
        am.move_type = 'in_invoice'
        AND am.state != 'cancel'
        AND {state_clause}

    ORDER BY
        effective_due_date NULLS LAST,
        am.invoice_date DESC NULLS LAST,
        am.id DESC,
        aml.sequence NULLS LAST
    """

    engine = db_engine()
    try:
        df = pd.read_sql_query(query, engine)

        def normalize_product_name(x):
            if isinstance(x, dict):
                return x.get('es_GT') or x.get('en_US') or str(x)
            return "" if x is None else str(x)

        if "product_name" in df.columns:
            df["product_name"] = df["product_name"].apply(normalize_product_name)

        if "bill_notes" in df.columns:
            df["bill_notes"] = df["bill_notes"].apply(clean_html_text)

        text_cols = [
            "bill_number", "vendor_reference", "invoice_origin", "supplier_name",
            "supplier_phone", "supplier_email", "bill_notes", "po_numbers",
            "line_label", "line_ref", "product_name", "product_code"
        ]
        for c in text_cols:
            if c in df.columns:
                df[c] = df[c].apply(clean_excel_value)

        return df
    finally:
        engine.dispose()

# -----------------------------
# Excel builder
# -----------------------------
def create_excel_report(df, output_path):
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    overdue_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    pending_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Invoice-level dataset
    header_cols = [
        "bill_id","bill_number","bill_state","payment_state","currency",
        "supplier_name","supplier_phone","supplier_email",
        "vendor_reference","invoice_origin","po_numbers",
        "invoice_date","invoice_date_due","effective_due_date",
        "bill_total_signed","bill_residual_signed","amount_paid_signed",
        "days_overdue","bill_notes"
    ]
    inv = df[header_cols].drop_duplicates(subset=["bill_id"]).copy()

    inv["effective_due_date"] = pd.to_datetime(inv["effective_due_date"], errors="coerce")

    inv_unpaid = inv[inv["bill_residual_signed"] > 0].copy()
    inv_paid = inv[inv["bill_residual_signed"] <= 0].copy()

    today = pd.Timestamp.today().normalize()
    next_7 = today + pd.Timedelta(days=7)

    inv_next7 = inv_unpaid[
        (inv_unpaid["effective_due_date"].dt.normalize() >= today) &
        (inv_unpaid["effective_due_date"].dt.normalize() <= next_7)
    ].copy()

    # -----------------
    # Sheet 1: Next 7 Days
    # -----------------
    ws1 = wb.create_sheet(safe_sheet_title("Next 7 Days (Budget)"))
    s_cols = [
        "effective_due_date","days_overdue","bill_number","supplier_name","currency",
        "bill_total_signed","amount_paid_signed","bill_residual_signed",
        "payment_state","bill_state","po_numbers","vendor_reference","invoice_origin","bill_notes"
    ]

    inv_next7 = inv_next7.sort_values(["effective_due_date","supplier_name","bill_residual_signed"], ascending=[True, True, False])

    for c_idx, col in enumerate(s_cols, 1):
        cell = ws1.cell(row=1, column=c_idx, value=col.replace("_"," ").title())
        cell.fill = header_fill
        cell.font = header_font

    for r_idx, row in enumerate(inv_next7.to_dict(orient="records"), start=2):
        for c_idx, col in enumerate(s_cols, 1):
            val = row.get(col)
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)

            if col in ["bill_total_signed","amount_paid_signed","bill_residual_signed"]:
                cell.number_format = '#,##0.00'
            if col == "effective_due_date" and pd.notna(val):
                cell.number_format = 'dd/mm/yyyy'
            if col == "bill_notes":
                cell.alignment = Alignment(wrap_text=True, vertical='top')

            if row.get("days_overdue", 0) and row.get("days_overdue", 0) > 0:
                cell.fill = overdue_fill
            elif col == "bill_residual_signed" and (row.get("bill_residual_signed", 0) > 0):
                cell.fill = pending_fill

    # -----------------
    # Sheet 2: All Unpaid Bills
    # -----------------
    ws2 = wb.create_sheet(safe_sheet_title("All Unpaid Bills"))
    inv_unpaid = inv_unpaid.sort_values(["days_overdue","effective_due_date","bill_residual_signed"], ascending=[False, True, False])

    for c_idx, col in enumerate(s_cols, 1):
        cell = ws2.cell(row=1, column=c_idx, value=col.replace("_"," ").title())
        cell.fill = header_fill
        cell.font = header_font

    for r_idx, row in enumerate(inv_unpaid.to_dict(orient="records"), start=2):
        for c_idx, col in enumerate(s_cols, 1):
            val = row.get(col)
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)

            if col in ["bill_total_signed","amount_paid_signed","bill_residual_signed"]:
                cell.number_format = '#,##0.00'
            if col == "effective_due_date" and pd.notna(val):
                cell.number_format = 'dd/mm/yyyy'
            if col == "bill_notes":
                cell.alignment = Alignment(wrap_text=True, vertical='top')

            if row.get("days_overdue", 0) and row.get("days_overdue", 0) > 0:
                cell.fill = overdue_fill

    # -----------------
    # Sheet 3: Lines Detail
    # -----------------
    ws3 = wb.create_sheet(safe_sheet_title("Bill Lines (Detail)"))

    # real lines only
    lines = df[df["display_type"].isna()].copy()
    lines["effective_due_date"] = pd.to_datetime(lines["effective_due_date"], errors="coerce")

    s3_cols = [
        "bill_number","supplier_name","currency","effective_due_date","payment_state","bill_state",
        "po_numbers","vendor_reference","invoice_origin",
        "product_code","product_name","line_label","quantity","price_unit","price_subtotal","price_total"
    ]

    for c_idx, col in enumerate(s3_cols, 1):
        cell = ws3.cell(row=1, column=c_idx, value=col.replace("_"," ").title())
        cell.fill = header_fill
        cell.font = header_font

    for r_idx, row in enumerate(lines.to_dict(orient="records"), start=2):
        for c_idx, col in enumerate(s3_cols, 1):
            val = row.get(col)
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)

            if col in ["price_unit","price_subtotal","price_total"]:
                cell.number_format = '#,##0.00'
            if col == "effective_due_date" and pd.notna(val):
                cell.number_format = 'dd/mm/yyyy'
            if col == "line_label":
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    # -----------------
    # Sheet 4: Paid History
    # -----------------
    ws4 = wb.create_sheet(safe_sheet_title("Paid History"))
    s4_cols = [
        "invoice_date","bill_number","supplier_name","currency",
        "bill_total_signed","amount_paid_signed","bill_residual_signed",
        "payment_state","bill_state","po_numbers","vendor_reference","invoice_origin","bill_notes"
    ]

    inv_paid = inv_paid.sort_values(["invoice_date","bill_number"], ascending=[False, False])

    for c_idx, col in enumerate(s4_cols, 1):
        cell = ws4.cell(row=1, column=c_idx, value=col.replace("_"," ").title())
        cell.fill = header_fill
        cell.font = header_font

    for r_idx, row in enumerate(inv_paid.to_dict(orient="records"), start=2):
        for c_idx, col in enumerate(s4_cols, 1):
            val = row.get(col)
            cell = ws4.cell(row=r_idx, column=c_idx, value=val)

            if col in ["bill_total_signed","amount_paid_signed","bill_residual_signed"]:
                cell.number_format = '#,##0.00'
            if col == "invoice_date" and pd.notna(val):
                cell.number_format = 'dd/mm/yyyy'
            if col == "bill_notes":
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    # -----------------
    # Formatting: filters + widths
    # -----------------
    for ws in wb.worksheets:
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col_idx).value or ""
            col_letter = ws.cell(row=1, column=col_idx).column_letter

            if any(k in str(header) for k in ["Notes", "Label"]):
                ws.column_dimensions[col_letter].width = 55
            else:
                max_len = 0
                for r in range(1, min(ws.max_row, 150) + 1):
                    v = ws.cell(row=r, column=col_idx).value
                    if v is not None:
                        max_len = max(max_len, len(str(v)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    wb.save(output_path)
    print(f"Report saved to: {output_path}")

# -----------------------------
# Summary
# -----------------------------
def print_summary(df):
    header_cols = [
        "bill_id","bill_number","bill_state","payment_state","currency",
        "supplier_name","effective_due_date",
        "bill_total_signed","bill_residual_signed","amount_paid_signed","days_overdue"
    ]
    inv = df[header_cols].drop_duplicates(subset=["bill_id"]).copy()

    unpaid = inv[inv["bill_residual_signed"] > 0]
    paid = inv[inv["bill_residual_signed"] <= 0]

    print("\n" + "="*60)
    print("VENDOR BILLS SUMMARY")
    print("="*60)
    print(f"Total bills (not canceled): {inv['bill_id'].nunique()}")
    print(f"Unpaid/partial (residual > 0): {unpaid['bill_id'].nunique()}")
    print(f"Paid (residual <= 0): {paid['bill_id'].nunique()}")

    # quick diagnostic
    if inv["bill_residual_signed"].max() == 0:
        print("\n⚠ DIAGNOSTIC: amount_residual_signed is 0 for ALL fetched bills.")
        print("   - If this is unexpected, confirm you are querying the correct DB/company.")
        print("   - If bills are posted and still 0, check if your Odoo uses a different residual field/customization.")

    for cur in sorted(unpaid["currency"].dropna().unique()):
        tot = unpaid[unpaid["currency"] == cur]["bill_residual_signed"].sum()
        sym = "$" if cur == "USD" else "Q" if cur == "GTQ" else cur
        print(f"  - Pending {cur}: {sym}{tot:,.2f}")

    overdue = unpaid[unpaid["days_overdue"] > 0]
    if len(overdue):
        print(f"Overdue bills: {len(overdue)} | Avg days overdue: {overdue['days_overdue'].mean():.1f}")

    print("="*60)

# -----------------------------
# Main
# -----------------------------
def main():
    print("VENDOR BILLS + PO LINK REPORT")
    print(f"Started at: {datetime.datetime.now()}")
    print("-"*60)

    try:
        date_folder = datetime.datetime.now().strftime('%Y-%m-%d')
        output_dir = os.path.join(BASE_DIR, date_folder)
        os.makedirs(output_dir, exist_ok=True)

        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Vendor_Bills_PO_Report_{timestamp}.xlsx'
        output_path = os.path.join(output_dir, filename)

        # IMPORTANT: only_posted=True is recommended
        df = get_vendor_bill_data(only_posted=True)

        if df.empty:
            print("No vendor bills found.")
            return

        print_summary(df)

        print("\nGenerating Excel report...")
        create_excel_report(df, output_path)

        print("\n✓ COMPLETE! Report generated successfully.")
        print("Tabs:")
        print("  1) Next 7 Days (Budget)")
        print("  2) All Unpaid Bills")
        print("  3) Bill Lines (Detail)")
        print("  4) Paid History")

    except Exception as e:
        print(f"\n✗ ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
