import sys
sys.path.append('/opt/torelo-kpi/scripts')
from config import DATABASE_CONFIG, BASE_DIR, WEB_DIR
db_params = DATABASE_CONFIG
import pandas as pd
from sqlalchemy import create_engine
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import os
import re

# =============================================
# CONFIGURATION SECTION - ADJUST AS NEEDED
# =============================================

# Database connection parameters
db_params = {
    "dbname": "master",
    "user": "rudy",
    "password": "Kaibil10!z",
    "host": "3.18.71.107",
    "port": "5432"
}

# Currency conversion rate
USD_TO_GTQ_RATE = 7.7

# Torelo URL Configuration
TORELO_BASE_URL = "https://www.torelo.net/web#"
DEFAULT_CIDS = "1"  # Company ID

# Menu and Action IDs
MENU_IDS = {
    "project": "385",
    "task": "385"
}

ACTIONS = {
    "project_kanban": "139",
    "project_form": "139",
    "task_kanban": "557",
    "task_form": "557"
}

USE_ALTERNATIVE_URL_FORMAT = False

# =============================================

# UPDATED QUERY WITH FECHA_INICIO, PURCHASE ORDER DATA, AND CURRENCY INFO
# Using f-string to properly inject the conversion rate
query_project_plan = f"""
WITH task_hierarchy AS (
    SELECT 
        -- Project Information
        proj.id AS project_id,
        proj.name ->> 'es_GT' AS project_name,
        proj.description AS project_description,
        proj.date_start AS project_start_date,
        proj.date AS project_end_date,
        
        -- Task Information
        task.id AS task_id,
        task.name AS task_name,
        REGEXP_REPLACE(task.description, '<[^>]+>', '', 'g') AS task_description,
        task.parent_id,
        task.sequence,
        
        -- UPDATED: Use ONLY fecha_inicio field - no fallback
        task.fecha_inicio AS task_start_date,
        task.date_deadline AS task_end_date,
        task.date_end AS actual_end_date,
        
        -- Status
        stage.name ->> 'es_GT' AS status,
        task.is_closed,
        task.progress AS progress_percentage,
        
        -- Hours
        task.planned_hours,
        task.effective_hours,
        task.remaining_hours,
        
        -- Assignment
        STRING_AGG(DISTINCT partner.name, ', ') AS assigned_to,
        
        -- Level calculation for hierarchy
        CASE 
            WHEN task.parent_id IS NULL THEN 0
            ELSE 1
        END AS task_level,
        
        -- Active flags
        proj.active AS project_active,
        task.active AS task_active
        
    FROM project_task task
    LEFT JOIN project_project proj ON task.project_id = proj.id
    LEFT JOIN project_task_type stage ON task.stage_id = stage.id
    LEFT JOIN project_task_user_rel rel ON task.id = rel.task_id
    LEFT JOIN res_users usr ON rel.user_id = usr.id
    LEFT JOIN res_partner partner ON usr.partner_id = partner.id
    
    WHERE proj.active = true AND task.active = true
    
    GROUP BY 
        proj.id, proj.name, proj.description, proj.date_start, proj.date,
        task.id, task.name, task.description, task.parent_id, task.sequence,
        task.fecha_inicio, task.date_deadline, task.date_end,
        stage.name, task.is_closed, task.progress,
        task.planned_hours, task.effective_hours, task.remaining_hours,
        proj.active, task.active
),
dependencies AS (
    SELECT 
        t1.id AS task_id,
        STRING_AGG(t2.name || ' (' || dep.depends_on_id::text || ')', ', ') AS dependencies,
        STRING_AGG(dep.depends_on_id::text, ',') AS dependency_ids
    FROM task_dependencies_rel dep
    JOIN project_task t1 ON dep.task_id = t1.id
    JOIN project_task t2 ON dep.depends_on_id = t2.id
    WHERE t1.active = true AND t2.active = true
    GROUP BY t1.id
),
-- CORRECTED: Purchase Orders with proper pending calculation and no duplication
purchase_orders AS (
    WITH po_line_pending AS (
        -- Calculate pending at line level WITHOUT joining invoices to avoid duplication
        SELECT 
            task_po.task_id,
            po.id AS po_id,
            po.name AS po_number,
            po.state AS po_state,
            po.amount_total AS po_total,
            
            -- Currency
            CASE 
                WHEN po.currency_id = 163 THEN 'GTQ'
                WHEN po.currency_id = 2 THEN 'USD'
                ELSE rc.name
            END AS currency,
            
            -- Sum pending across all lines for this PO
            SUM(
                CASE 
                    WHEN pol.qty_invoiced = 0 OR pol.qty_invoiced IS NULL THEN 
                        pol.price_total
                    WHEN pol.qty_invoiced < pol.product_qty THEN 
                        (pol.product_qty - pol.qty_invoiced) * pol.price_unit
                    ELSE 
                        -- For fully invoiced lines, check if there's unpaid amount using subquery
                        COALESCE(
                            (SELECT MAX(ABS(am.amount_residual_signed))
                             FROM account_move_line aml
                             JOIN account_move am ON am.id = aml.move_id
                             WHERE aml.purchase_line_id = pol.id
                             AND am.move_type = 'in_invoice'
                             AND am.state != 'cancel'
                             AND am.payment_state != 'paid'
                             LIMIT 1), 
                            0
                        )
                END
            ) AS total_pending_original
            
        FROM project_task_purchase_order_rel task_po
        JOIN purchase_order po ON task_po.purchase_order_id = po.id
        JOIN purchase_order_line pol ON po.id = pol.order_id
        LEFT JOIN res_currency rc ON po.currency_id = rc.id
        WHERE po.state != 'cancel'
        GROUP BY task_po.task_id, po.id, po.name, po.state, po.amount_total, po.currency_id, rc.name
    )
    SELECT 
        task_id,
        COUNT(DISTINCT po_id) AS po_count,
        
        -- Convert to GTQ
        SUM(
            CASE 
                WHEN currency = 'USD' THEN 
                    total_pending_original * {USD_TO_GTQ_RATE}
                ELSE 
                    total_pending_original
            END
        ) AS total_po_amount_gtq,
        
        -- Create reference string with original totals
        STRING_AGG(
            DISTINCT 
            CASE 
                WHEN po_state = 'draft' THEN 
                    CASE 
                        WHEN currency = 'USD' THEN 
                            'RFQ-' || po_number || ' ($' || ROUND(po_total::numeric, 2)::text || ')'
                        ELSE 
                            'RFQ-' || po_number
                    END
                ELSE 
                    CASE 
                        WHEN currency = 'USD' THEN 
                            po_number || ' ($' || ROUND(po_total::numeric, 2)::text || ')'
                        ELSE 
                            po_number
                    END
            END, 
            ', '
        ) AS po_references,
        
        STRING_AGG(DISTINCT currency, ',') AS currencies_used
        
    FROM po_line_pending
    GROUP BY task_id
)
SELECT 
    th.*,
    dep.dependencies,
    dep.dependency_ids,
    -- Add PO information with pending amounts
    COALESCE(po.po_count, 0) AS po_count,
    COALESCE(po.total_po_amount_gtq, 0) AS total_po_amount,
    po.po_references,
    po.currencies_used
FROM task_hierarchy th
LEFT JOIN dependencies dep ON th.task_id = dep.task_id
LEFT JOIN purchase_orders po ON th.task_id = po.task_id
ORDER BY 
    th.project_name,
    COALESCE(th.parent_id, th.task_id),
    th.sequence,
    th.task_id;
"""

def create_torelo_url(entity_id, is_project=False, view_type='form', project_id=None):
    """Create a Torelo URL for a project or task"""
    if not entity_id or str(entity_id).startswith('P'):
        entity_id = str(entity_id).replace('P', '') if entity_id else '0'
    
    try:
        entity_id = int(entity_id)
    except (ValueError, TypeError):
        print(f"Warning: Invalid entity ID: {entity_id}")
        entity_id = 0
    
    if is_project:
        model = 'project.project'
        action = ACTIONS.get("project_form", "139")
        menu_id = MENU_IDS.get("project", "385")
        return f"{TORELO_BASE_URL}id={entity_id}&cids={DEFAULT_CIDS}&menu_id={menu_id}&action={action}&active_id={entity_id}&model={model}&view_type={view_type}"
    else:
        model = 'project.task'
        action = ACTIONS.get("task_form", "557")
        menu_id = MENU_IDS.get("task", "385")
        
        if not project_id:
            print(f"Warning: No project_id provided for task {entity_id}")
            project_id = entity_id
        
        return f"{TORELO_BASE_URL}id={entity_id}&cids={DEFAULT_CIDS}&menu_id={menu_id}&action={action}&active_id={project_id}&model={model}&view_type={view_type}"

def clean_html_text(text):
    """Remove HTML tags and clean text"""
    if pd.isna(text) or text is None:
        return ""
    clean_text = re.sub('<.*?>', '', str(text))
    clean_text = ' '.join(clean_text.split())
    clean_text = clean_text.replace('&nbsp;', ' ')
    clean_text = clean_text.replace('&amp;', '&')
    clean_text = clean_text.replace('&lt;', '<')
    clean_text = clean_text.replace('&gt;', '>')
    clean_text = clean_text.replace('&quot;', '"')
    clean_text = clean_text.replace('\n', ' ')
    clean_text = clean_text.replace('\r', ' ')
    if len(clean_text) > 200:
        clean_text = clean_text[:197] + "..."
    return clean_text.strip()

def format_currency_gtq(value):
    """Format currency value in Quetzales"""
    if pd.isna(value) or value == 0:
        return ""
    return f"Q{value:,.2f}"

def fetch_project_data(query, db_params):
    """Fetch project plan data from database"""
    conn_str = f"postgresql://{db_params['user']}:{db_params['password']}@{db_params['host']}:{db_params['port']}/{db_params['dbname']}"
    engine = create_engine(conn_str)
    print("Connecting to database for project plan view...")
    
    try:
        df = pd.read_sql_query(query, engine)
        print(f"Successfully fetched {len(df)} task records")
        
        if len(df) > 0:
            print("\nSample Project IDs found:")
            sample_projects = df[['project_id', 'project_name']].drop_duplicates().head(5)
            for _, row in sample_projects.iterrows():
                print(f"  - ID: {row['project_id']}, Name: {row['project_name']}")
            
            print("\nSample Task IDs with PO info:")
            sample_tasks = df[['task_id', 'task_name', 'po_count', 'total_po_amount', 'po_references']].head(5)
            for _, row in sample_tasks.iterrows():
                po_info = f"POs: {row['po_count']}, Total: Q{row['total_po_amount']:,.2f}" if row['po_count'] > 0 else "No POs"
                po_nums = f" ({row['po_references']})" if pd.notna(row['po_references']) else ""
                print(f"  - ID: {row['task_id']}, Name: {row['task_name']}, {po_info}{po_nums}")
        
        # Ensure date columns are properly parsed
        date_columns = ['project_start_date', 'project_end_date', 'task_start_date', 'task_end_date', 'actual_end_date']
        for col in date_columns:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce')
        
        # Ensure status column has no null values
        if 'status' in df.columns:
            df['status'] = df['status'].fillna('')
        
        return df
    except Exception as e:
        print(f"Error fetching data: {e}")
        raise
    finally:
        engine.dispose()

def get_project_sort_order(df):
    """Get project sort order based on earliest task end date"""
    project_dates = {}
    
    for project_id in df['project_id'].unique():
        project_tasks = df[df['project_id'] == project_id]
        # Get the earliest task end date for this project
        valid_dates = project_tasks[project_tasks['task_end_date'].notna()]['task_end_date']
        if len(valid_dates) > 0:
            project_dates[project_id] = valid_dates.min()
        else:
            # If no valid dates, use a far future date
            project_dates[project_id] = pd.Timestamp('2099-12-31')
    
    # Sort projects by their earliest task date
    sorted_projects = sorted(project_dates.items(), key=lambda x: x[1])
    return [pid for pid, _ in sorted_projects]

def create_open_projects_hierarchy(df):
    """Create hierarchy for open/active projects only (excluding completed) with task sorting by end date"""
    completed_statuses = ['Hecho', 'hecho', 'Completo', 'completo', 'Done', 'done', 
                         'Completado', 'completado', 'Finalizado', 'finalizado']
    
    df = df.copy()
    df['status'] = df['status'].fillna('')
    
    df_open = df[~df['status'].isin(completed_statuses)].copy()
    
    # Get sorted project order based on earliest task dates
    project_order = get_project_sort_order(df_open)
    
    hierarchy_data = []
    
    # Updated column order with renamed columns
    columns_order = ['Level', 'ID', 'Task Name', 'Descripción', 'Fecha Inicio', 
                    'End Date', 'Días', 'Restante', 'Assigned To', 'Dependencies', 
                    'Status', 'PO Count', 'Pendiente (Q)', 'PO/RFQ Numbers', 'Type', 'Has End Date', 'Torelo Link']
    
    for project_id in project_order:
        project_info = df[df['project_id'] == project_id].iloc[0]
        project_name = project_info['project_name'] if pd.notna(project_info['project_name']) else f"Project {project_id}"
        
        project_tasks = df_open[df_open['project_id'] == project_id].copy()
        
        try:
            project_start_date = pd.to_datetime(project_info['project_start_date'])
            project_end_date = pd.to_datetime(project_info['project_end_date'])
            
            if pd.notna(project_start_date) and pd.notna(project_end_date):
                project_duration = (project_end_date - project_start_date).days + 1
            else:
                project_duration = ''
        except Exception as e:
            project_duration = ''
            print(f"Error calculating duration for project {project_name}: {e}")
        
        # Calculate total PO amount for entire project
        project_total_po = project_tasks['total_po_amount'].sum()
        
        # Add project header
        hierarchy_data.append({
            'Level': 0,
            'ID': f"P{project_id}",
            'Task Name': project_name.upper(),
            'Descripción': clean_html_text(project_info['project_description']),
            'Fecha Inicio': project_info['project_start_date'],
            'End Date': project_info['project_end_date'],
            'Días': project_duration,
            'Restante': '',
            'Assigned To': '',
            'Dependencies': '',
            'Status': 'PROJECT',
            'PO Count': '',
            'Pendiente (Q)': format_currency_gtq(project_total_po),
            'PO/RFQ Numbers': '',
            'Type': 'Project',
            'Has End Date': True,
            'Torelo Link': create_torelo_url(project_id, is_project=True)
        })
        
        # Calculate days until due for each task
        today = pd.Timestamp.now()
        for idx, row in project_tasks.iterrows():
            if pd.notna(row['task_end_date']):
                project_tasks.at[idx, 'days_until_due'] = (pd.to_datetime(row['task_end_date']) - today).days
            else:
                project_tasks.at[idx, 'days_until_due'] = None
        
        project_tasks['is_overdue'] = False
        for idx, row in project_tasks.iterrows():
            if pd.notna(row['days_until_due']) and row['days_until_due'] < 0 and row['is_closed'] == False:
                project_tasks.at[idx, 'is_overdue'] = True
        
        # Sort parent tasks by end date
        parent_tasks = project_tasks[project_tasks['parent_id'].isna()].copy()
        parent_tasks['sort_date'] = parent_tasks['task_end_date'].fillna(pd.Timestamp('2099-12-31'))
        parent_tasks = parent_tasks.sort_values('sort_date')
        
        for _, task in parent_tasks.iterrows():
            task_name = task['task_name'] if pd.notna(task['task_name']) else f"Task {task['task_id']}"
            task_description = clean_html_text(task['task_description'])
            days_until = task.get('days_until_due', None)
            has_end_date = pd.notna(task['task_end_date'])
            
            task_start = pd.to_datetime(task['task_start_date'])
            task_end = pd.to_datetime(task['task_end_date'])
            if pd.notna(task_start) and pd.notna(task_end):
                task_duration = (task_end - task_start).days + 1
            else:
                task_duration = ''
            
            if days_until is not None and pd.notna(days_until) and not pd.isna(days_until):
                try:
                    days_until = int(days_until)
                except:
                    days_until = ''
            else:
                days_until = ''
            
            task_url = create_torelo_url(task['task_id'], is_project=False, project_id=project_id)
            
            hierarchy_data.append({
                'Level': 1,
                'ID': str(task['task_id']),
                'Task Name': task_name,
                'Descripción': task_description,
                'Fecha Inicio': task['task_start_date'],
                'End Date': task['task_end_date'],
                'Días': task_duration,
                'Restante': days_until if days_until is not None else '',
                'Assigned To': task['assigned_to'] if pd.notna(task['assigned_to']) else '',
                'Dependencies': task['dependencies'] if pd.notna(task['dependencies']) else '',
                'Status': task['status'] if pd.notna(task['status']) else '',
                'PO Count': int(task['po_count']) if task['po_count'] > 0 else '',
                'Pendiente (Q)': format_currency_gtq(task['total_po_amount']),
                'PO/RFQ Numbers': task['po_references'] if pd.notna(task['po_references']) else '',
                'Type': 'Task',
                'Has End Date': has_end_date,
                'Torelo Link': task_url
            })
            
            # Sort subtasks by end date
            subtasks = project_tasks[project_tasks['parent_id'] == task['task_id']].copy()
            subtasks['sort_date'] = subtasks['task_end_date'].fillna(pd.Timestamp('2099-12-31'))
            subtasks = subtasks.sort_values('sort_date')
            
            for _, subtask in subtasks.iterrows():
                subtask_name = subtask['task_name'] if pd.notna(subtask['task_name']) else f"Subtask {subtask['task_id']}"
                subtask_description = clean_html_text(subtask['task_description'])
                days_until = subtask.get('days_until_due', None)
                has_end_date = pd.notna(subtask['task_end_date'])
                
                subtask_start = pd.to_datetime(subtask['task_start_date'])
                subtask_end = pd.to_datetime(subtask['task_end_date'])
                if pd.notna(subtask_start) and pd.notna(subtask_end):
                    subtask_duration = (subtask_end - subtask_start).days + 1
                else:
                    subtask_duration = ''
                
                if days_until is not None and pd.notna(days_until) and not pd.isna(days_until):
                    try:
                        days_until = int(days_until)
                    except:
                        days_until = ''
                else:
                    days_until = ''
                
                hierarchy_data.append({
                    'Level': 2,
                    'ID': str(subtask['task_id']),
                    'Task Name': f"  → {subtask_name}",
                    'Descripción': subtask_description,
                    'Fecha Inicio': subtask['task_start_date'],
                    'End Date': subtask['task_end_date'],
                    'Días': subtask_duration,
                    'Restante': days_until if days_until is not None else '',
                    'Assigned To': subtask['assigned_to'] if pd.notna(subtask['assigned_to']) else '',
                    'Dependencies': subtask['dependencies'] if pd.notna(subtask['dependencies']) else '',
                    'Status': subtask['status'] if pd.notna(subtask['status']) else '',
                    'PO Count': int(subtask['po_count']) if subtask['po_count'] > 0 else '',
                    'Pendiente (Q)': format_currency_gtq(subtask['total_po_amount']),
                    'PO/RFQ Numbers': subtask['po_references'] if pd.notna(subtask['po_references']) else '',
                    'Type': 'Subtask',
                    'Has End Date': has_end_date,
                    'Torelo Link': create_torelo_url(subtask['task_id'], is_project=False, project_id=project_id)
                })
    
    df_result = pd.DataFrame(hierarchy_data)
    return df_result[columns_order] if len(df_result) > 0 else pd.DataFrame(columns=columns_order)

def create_project_hierarchy(df):
    """Create a hierarchical structure of ALL projects and tasks with task sorting"""
    # Get sorted project order
    project_order = get_project_sort_order(df)
    
    hierarchy_data = []
    
    for project_id in project_order:
        project_info = df[df['project_id'] == project_id].iloc[0]
        project_name = project_info['project_name'] if pd.notna(project_info['project_name']) else f"Project {project_id}"
        
        project_tasks = df[df['project_id'] == project_id].copy()
        
        try:
            project_start_date = pd.to_datetime(project_info['project_start_date'])
            project_end_date = pd.to_datetime(project_info['project_end_date'])
            
            if pd.notna(project_start_date) and pd.notna(project_end_date):
                project_duration = (project_end_date - project_start_date).days + 1
            else:
                project_duration = ''
        except Exception as e:
            project_duration = ''
            print(f"Error calculating duration for project {project_name}: {e}")
        
        # Calculate total PO amount for entire project
        project_total_po = project_tasks['total_po_amount'].sum()
        
        # Add project header with Torelo Link
        hierarchy_data.append({
            'Level': 0,
            'ID': f"P{project_id}",
            'Task Name': project_name.upper(),
            'Descripción': clean_html_text(project_info['project_description']),
            'Fecha Inicio': project_info['project_start_date'],
            'End Date': project_info['project_end_date'],
            'Días': project_duration,
            'Restante': '',
            'Assigned To': '',
            'Dependencies': '',
            'Status': 'PROJECT',
            'PO Count': '',
            'Pendiente (Q)': format_currency_gtq(project_total_po),
            'PO/RFQ Numbers': '',
            'Type': 'Project',
            'Has End Date': True,
            'Torelo Link': create_torelo_url(project_id, is_project=True)
        })
        
        # Calculate days until due for each task
        today = pd.Timestamp.now()
        for idx, row in project_tasks.iterrows():
            if pd.notna(row['task_end_date']):
                project_tasks.at[idx, 'days_until_due'] = (pd.to_datetime(row['task_end_date']) - today).days
            else:
                project_tasks.at[idx, 'days_until_due'] = None
        
        project_tasks['is_overdue'] = False
        for idx, row in project_tasks.iterrows():
            if pd.notna(row['days_until_due']) and row['days_until_due'] < 0 and row['is_closed'] == False:
                project_tasks.at[idx, 'is_overdue'] = True
        
        # Sort parent tasks by end date
        parent_tasks = project_tasks[project_tasks['parent_id'].isna()].copy()
        parent_tasks['sort_date'] = parent_tasks['task_end_date'].fillna(pd.Timestamp('2099-12-31'))
        parent_tasks = parent_tasks.sort_values('sort_date')
        
        for _, task in parent_tasks.iterrows():
            task_name = task['task_name'] if pd.notna(task['task_name']) else f"Task {task['task_id']}"
            task_description = clean_html_text(task['task_description'])
            days_until = task.get('days_until_due', None)
            has_end_date = pd.notna(task['task_end_date'])
            
            task_start = pd.to_datetime(task['task_start_date'])
            task_end = pd.to_datetime(task['task_end_date'])
            if pd.notna(task_start) and pd.notna(task_end):
                task_duration = (task_end - task_start).days
            else:
                task_duration = ''
            
            hierarchy_data.append({
                'Level': 1,
                'ID': str(task['task_id']),
                'Task Name': task_name,
                'Descripción': task_description,
                'Fecha Inicio': task['task_start_date'],
                'End Date': task['task_end_date'],
                'Días': task_duration,
                'Restante': days_until if days_until is not None else '',
                'Assigned To': task['assigned_to'] if pd.notna(task['assigned_to']) else '',
                'Dependencies': task['dependencies'] if pd.notna(task['dependencies']) else '',
                'Status': task['status'] if pd.notna(task['status']) else '',
                'PO Count': int(task['po_count']) if task['po_count'] > 0 else '',
                'Pendiente (Q)': format_currency_gtq(task['total_po_amount']),
                'PO/RFQ Numbers': task['po_references'] if pd.notna(task['po_references']) else '',
                'Type': 'Task',
                'Has End Date': has_end_date,
                'Torelo Link': create_torelo_url(task['task_id'], is_project=False, project_id=project_id)
            })
            
            # Sort subtasks by end date
            subtasks = project_tasks[project_tasks['parent_id'] == task['task_id']].copy()
            subtasks['sort_date'] = subtasks['task_end_date'].fillna(pd.Timestamp('2099-12-31'))
            subtasks = subtasks.sort_values('sort_date')
            
            for _, subtask in subtasks.iterrows():
                subtask_name = subtask['task_name'] if pd.notna(subtask['task_name']) else f"Subtask {subtask['task_id']}"
                subtask_description = clean_html_text(subtask['task_description'])
                days_until = subtask.get('days_until_due', None)
                has_end_date = pd.notna(subtask['task_end_date'])
                
                subtask_start = pd.to_datetime(subtask['task_start_date'])
                subtask_end = pd.to_datetime(subtask['task_end_date'])
                if pd.notna(subtask_start) and pd.notna(subtask_end):
                    subtask_duration = (subtask_end - subtask_start).days
                else:
                    subtask_duration = ''
                
                hierarchy_data.append({
                    'Level': 2,
                    'ID': str(subtask['task_id']),
                    'Task Name': f"  → {subtask_name}",
                    'Descripción': subtask_description,
                    'Fecha Inicio': subtask['task_start_date'],
                    'End Date': subtask['task_end_date'],
                    'Días': subtask_duration,
                    'Restante': days_until if days_until is not None else '',
                    'Assigned To': subtask['assigned_to'] if pd.notna(subtask['assigned_to']) else '',
                    'Dependencies': subtask['dependencies'] if pd.notna(subtask['dependencies']) else '',
                    'Status': subtask['status'] if pd.notna(subtask['status']) else '',
                    'PO Count': int(subtask['po_count']) if subtask['po_count'] > 0 else '',
                    'Pendiente (Q)': format_currency_gtq(subtask['total_po_amount']),
                    'PO/RFQ Numbers': subtask['po_references'] if pd.notna(subtask['po_references']) else '',
                    'Type': 'Subtask',
                    'Has End Date': has_end_date,
                    'Torelo Link': create_torelo_url(subtask['task_id'], is_project=False, project_id=project_id)
                })
    
    df_result = pd.DataFrame(hierarchy_data)
    columns_order = ['Level', 'ID', 'Task Name', 'Descripción', 'Fecha Inicio', 
                    'End Date', 'Días', 'Restante', 'Assigned To', 'Dependencies', 
                    'Status', 'PO Count', 'Pendiente (Q)', 'PO/RFQ Numbers', 'Type', 'Has End Date', 'Torelo Link']
    return df_result[columns_order] if len(df_result) > 0 else pd.DataFrame(columns=columns_order)

def format_project_plan_excel(writer, df_plan, sheet_name='Project Plan', add_status_colors=False):
    """Format the Excel sheet to look like MS Project with PO information"""
    # Remove Has End Date, Torelo Link, and Restante columns before writing to Excel
    df_display = df_plan.drop(columns=['Has End Date', 'Torelo Link', 'Restante'], errors='ignore')
    df_display.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.sheets[sheet_name]
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    project_font = Font(bold=True, size=12, color="FFFFFF")
    project_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    task_font = Font(bold=True, size=10)
    subtask_font = Font(size=10, italic=True)
    
    # Style for hyperlinks
    link_font = Font(color="0563C1", underline="single")
    
    # Style for PO amounts (now in Quetzales)
    po_font = Font(bold=True, color="008000")  # Green for money
    
    # Color fills for status
    overdue_fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
    blocked_fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
    due_soon_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    no_end_date_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format headers
    for col_num, column_title in enumerate(df_display.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Get column indices
    status_col_idx = None
    end_date_col_idx = None
    type_col_idx = None
    id_col_idx = None
    po_amount_col_idx = None
    po_numbers_col_idx = None
    
    if 'Status' in df_display.columns:
        status_col_idx = df_display.columns.get_loc('Status') + 1
    if 'End Date' in df_display.columns:
        end_date_col_idx = df_display.columns.get_loc('End Date') + 1
    if 'Type' in df_display.columns:
        type_col_idx = df_display.columns.get_loc('Type') + 1
    if 'ID' in df_display.columns:
        id_col_idx = df_display.columns.get_loc('ID') + 1
    if 'Pendiente (Q)' in df_display.columns:
        po_amount_col_idx = df_display.columns.get_loc('Pendiente (Q)') + 1
    if 'PO/RFQ Numbers' in df_display.columns:
        po_numbers_col_idx = df_display.columns.get_loc('PO/RFQ Numbers') + 1
    
    # Format data rows
    for row_num in range(2, len(df_plan) + 2):
        row_data = df_plan.iloc[row_num - 2]
        row_type = row_data['Type']
        
        dias_col_idx = None
        if 'Días' in df_display.columns:
            dias_col_idx = df_display.columns.get_loc('Días') + 1
        
        for col_num in range(1, len(df_display.columns) + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.border = border
            
            # Format Type column with hyperlink
            if type_col_idx and col_num == type_col_idx:
                if row_type == 'Task':
                    url = row_data['Torelo Link']
                    cell.value = "Task"
                    cell.hyperlink = url
                    cell.font = link_font
                    cell.alignment = Alignment(horizontal="center")
                    continue
                elif row_type == 'Subtask':
                    url = row_data['Torelo Link']
                    cell.value = "Subtask"
                    cell.hyperlink = url
                    cell.font = Font(color="0563C1", underline="single", italic=True)
                    cell.alignment = Alignment(horizontal="center")
                    continue
                elif row_type == 'Project':
                    url = row_data['Torelo Link']
                    cell.value = "Project"
                    cell.hyperlink = url
                    cell.font = Font(color="FFFFFF", underline="single", bold=True, size=12)
                    cell.alignment = Alignment(horizontal="center")
                    cell.fill = project_fill
                    continue
            
            # Format Pendiente (Q) column with green bold font
            if po_amount_col_idx and col_num == po_amount_col_idx:
                if cell.value and cell.value != "" and cell.value != "Q0.00":
                    cell.font = po_font
                    cell.alignment = Alignment(horizontal="right")
            
            # Center align numeric columns
            if col_num in [dias_col_idx]:
                cell.alignment = Alignment(horizontal="center")
            
            if row_type == 'Project':
                cell.font = project_font
                cell.fill = project_fill
            elif row_type == 'Task':
                cell.font = task_font
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            else:  # Subtask
                cell.font = subtask_font
        
        # Apply status colors if requested
        if add_status_colors and row_type in ['Task', 'Subtask']:
            status_value = str(row_data.get('Status', '')).lower()
            days_until = row_data.get('Restante', None)
            is_overdue = days_until is not None and isinstance(days_until, (int, float)) and days_until < 0
            has_end_date = row_data.get('Has End Date', True)
            
            if not has_end_date and dias_col_idx:
                dias_cell = worksheet.cell(row=row_num, column=dias_col_idx)
                dias_cell.fill = no_end_date_fill
            
            if status_col_idx:
                if is_overdue:
                    status_cell = worksheet.cell(row=row_num, column=status_col_idx)
                    status_cell.fill = overdue_fill
                elif 'bloqueado' in status_value or 'blocked' in status_value:
                    status_cell = worksheet.cell(row=row_num, column=status_col_idx)
                    status_cell.fill = blocked_fill
                elif isinstance(days_until, (int, float)) and 0 <= days_until <= 3:
                    status_cell = worksheet.cell(row=row_num, column=status_col_idx)
                    status_cell.fill = due_soon_fill
    
    # Adjust column widths
    column_widths = {
        'Level': 8,
        'ID': 0,  # Hide ID column
        'Task Name': 50,
        'Descripción': 60,
        'Fecha Inicio': 15,
        'End Date': 12,
        'Días': 10,
        'Assigned To': 25,
        'Dependencies': 30,
        'Status': 15,
        'PO Count': 12,
        'Pendiente (Q)': 18,
        'PO/RFQ Numbers': 45,
        'Type': 10
    }
    
    for col_num, column_title in enumerate(df_display.columns, 1):
        column_letter = get_column_letter(col_num)
        if column_title in column_widths:
            worksheet.column_dimensions[column_letter].width = column_widths[column_title]
    
    # Hide the ID column
    if id_col_idx:
        id_letter = get_column_letter(id_col_idx)
        worksheet.column_dimensions[id_letter].hidden = True
    
    # Format dates
    date_columns = ['Fecha Inicio', 'End Date']
    for col_name in date_columns:
        if col_name in df_display.columns:
            col_idx = df_display.columns.get_loc(col_name) + 1
            for row in range(2, len(df_plan) + 2):
                cell = worksheet.cell(row=row, column=col_idx)
                if cell.value and isinstance(cell.value, datetime):
                    cell.number_format = 'YYYY-MM-DD'
    
    # Freeze panes
    worksheet.freeze_panes = 'C2'

def create_gantt_data_sheet(writer, df):
    """Create a sheet with data formatted for easy Gantt chart creation"""
    gantt_data = df[df['task_level'] >= 0].copy()
    
    gantt_data['project_name'] = gantt_data['project_name'].fillna('Proyecto sin nombre')
    gantt_data['task_name'] = gantt_data['task_name'].fillna('Tarea sin nombre')
    
    gantt_data['task_description'] = gantt_data['task_description'].apply(clean_html_text)
    
    gantt_data['calculated_duration'] = ''
    gantt_data['torelo_url'] = ''
    today = pd.Timestamp.now()
    
    for idx, row in gantt_data.iterrows():
        start_date = pd.to_datetime(row['task_start_date'])
        end_date = pd.to_datetime(row['task_end_date'])
        if pd.notna(start_date) and pd.notna(end_date):
            gantt_data.at[idx, 'calculated_duration'] = (end_date - start_date).days + 1
        
        gantt_data.at[idx, 'torelo_url'] = create_torelo_url(row['task_id'], is_project=False, project_id=row['project_id'])
    
    # Sort by task end date
    gantt_data['sort_date'] = gantt_data['task_end_date'].fillna(pd.Timestamp('2099-12-31'))
    gantt_data = gantt_data.sort_values(['project_name', 'sort_date'])
    
    # Select relevant columns for Gantt - WITH PO information in Quetzales
    gantt_columns = [
        'project_name', 'task_name', 'task_description', 'task_start_date', 'task_end_date',
        'calculated_duration', 'assigned_to', 'dependencies', 'po_count', 'total_po_amount', 
        'po_references', 'torelo_url'
    ]
    
    gantt_df = gantt_data[gantt_columns].copy()
    gantt_df.columns = [
        'Proyecto', 'Tarea', 'Descripción', 'Fecha Inicio', 'Fecha Fin', 
        'Días', 'Asignado A', 'Dependencias', 'PO Count', 'Pendiente (Q)', 
        'PO/RFQ Numbers', 'Enlace Torelo'
    ]
    
    gantt_df.to_excel(writer, sheet_name='Datos Gantt', index=False)
    
    # Format the Gantt data sheet
    worksheet = writer.sheets['Datos Gantt']
    
    # Set specific column widths for better readability
    column_widths = {
        'A': 30,  # Proyecto
        'B': 40,  # Tarea
        'C': 50,  # Descripción
        'D': 15,  # Fecha Inicio
        'E': 12,  # Fecha Fin
        'F': 10,  # Días
        'G': 25,  # Asignado A
        'H': 30,  # Dependencias
        'I': 12,  # PO Count
        'J': 18,  # Pendiente (Q)
        'K': 45,  # PO/RFQ Numbers
        'L': 15   # Enlace Torelo
    }
    
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width
    
    # Format header row
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    link_font = Font(color="0000FF", underline="single")
    po_font = Font(bold=True, color="008000")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num in range(1, 13):  # 12 columns total
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Format data cells
    for row in range(2, gantt_df.shape[0] + 2):
        for col in range(1, 13):
            cell = worksheet.cell(row=row, column=col)
            cell.border = border
            # Center align numeric columns
            if col in [6, 9]:  # Días and PO Count columns
                cell.alignment = Alignment(horizontal="center")
            # Format Pendiente (Q) column
            elif col == 10:  # Pendiente (Q) column
                if cell.value and cell.value != 0:
                    cell.value = format_currency_gtq(cell.value)
                    cell.font = po_font
                    cell.alignment = Alignment(horizontal="right")
            # Format hyperlinks in column L (12)
            elif col == 12:  # Enlace Torelo column
                url = gantt_df.iloc[row-2]['Enlace Torelo']
                cell.value = "Abrir"
                cell.hyperlink = url
                cell.font = link_font
                cell.alignment = Alignment(horizontal="center")
    
    # Format dates
    for row in range(2, gantt_df.shape[0] + 2):
        for col in [4, 5]:  # Fecha Inicio and Fecha Fin columns
            cell = worksheet.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, datetime):
                cell.number_format = 'YYYY-MM-DD'
    
    # Freeze panes to keep header visible
    worksheet.freeze_panes = 'A2'

def add_color_legend(writer, sheet_name, start_row, include_summary_colors=False):
    """Add a color legend to explain the status colors"""
    worksheet = writer.sheets[sheet_name]
    
    # Define styles for legend
    legend_font = Font(bold=True, size=11)
    overdue_fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
    due_soon_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    no_end_date_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    complete_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Add legend title
    worksheet.cell(row=start_row, column=1, value="LEYENDA DE COLORES:").font = legend_font
    
    # Add legend items
    legend_items = [
        ("Rojo Claro", "Status: Tareas Vencidas o Bloqueadas", overdue_fill),
        ("Amarillo Claro", "Status: Tareas que Vencen en 3 Días o Menos", due_soon_fill),
        ("Amarillo Claro", "Días: Tareas sin Fecha de Finalización", no_end_date_fill)
    ]
    
    if include_summary_colors:
        legend_items.append(("Verde Claro", "% Completado: Proyectos 100% Completados", complete_fill))
    
    for i, (color_name, description, fill) in enumerate(legend_items):
        row = start_row + i + 1
        # Color box
        color_cell = worksheet.cell(row=row, column=2, value=color_name)
        color_cell.fill = fill
        color_cell.border = border
        # Description
        worksheet.cell(row=row, column=3, value=description)
    
    # Adjust column widths for legend
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 45

def create_project_summary_sheet(writer, df):
    """Create a summary sheet with project metrics including PO information in Quetzales"""
    completed_statuses = ['Hecho', 'hecho', 'Completo', 'completo', 'Done', 'done', 
                         'Completado', 'completado', 'Finalizado', 'finalizado']
    
    project_summary = []
    
    # Get sorted project order
    project_order = get_project_sort_order(df)
    
    for project_id in project_order:
        project_data = df[df['project_id'] == project_id]
        project_info = project_data.iloc[0]
        
        project_name = project_info['project_name'] if pd.notna(project_info['project_name']) else f"Project {project_id}"
        
        total_tasks = len(project_data)
        project_data_status = project_data.copy()
        project_data_status['status'] = project_data_status['status'].fillna('')
        completed_tasks = len(project_data_status[project_data_status['status'].isin(completed_statuses)])
        total_planned_hours = project_data['planned_hours'].sum()
        total_effective_hours = project_data['effective_hours'].sum()
        
        # Calculate PO information
        total_po_amount = project_data['total_po_amount'].sum()
        # Get unique PO references for this project
        po_refs = []
        for refs in project_data[project_data['po_references'].notna()]['po_references']:
            po_refs.extend(refs.split(', '))
        unique_po_refs = ', '.join(sorted(set(po_refs))) if po_refs else ''
        
        progress_values = project_data['progress_percentage'].dropna()
        if len(progress_values) > 0:
            avg_progress = progress_values.mean()
        else:
            avg_progress = 0
        
        project_summary.append({
            'Nombre del Proyecto': project_name,
            'Fecha Inicio': project_info['project_start_date'],
            'Fecha Fin': project_info['project_end_date'],
            'Total Tareas': total_tasks,
            'Tareas Completadas': completed_tasks,
            '% Completado': f"{(completed_tasks/total_tasks*100):.1f}%" if total_tasks > 0 else "0%",
            'Horas Planificadas': round(total_planned_hours, 1) if pd.notna(total_planned_hours) else 0,
            'Horas Reales': round(total_effective_hours, 1) if pd.notna(total_effective_hours) else 0,
            'Progreso Promedio': avg_progress,
            'Pendiente (Q)': format_currency_gtq(total_po_amount),
            'Órdenes de Compra': unique_po_refs,
            'Enlace Proyecto': create_torelo_url(project_id, is_project=True)
        })
    
    summary_df = pd.DataFrame(project_summary)
    summary_df.to_excel(writer, sheet_name='Resumen de Proyectos', index=False)
    
    # Format summary sheet
    worksheet = writer.sheets['Resumen de Proyectos']
    worksheet.sheet_properties.tabColor = "FFFF99"
    
    # Apply formatting
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    link_font = Font(color="0000FF", underline="single")
    po_font = Font(bold=True, color="008000")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num in range(1, len(summary_df.columns) + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Apply borders and formatting to data cells
    for row in range(2, len(summary_df) + 2):
        percentage_value = summary_df.iloc[row-2]['% Completado']
        is_complete = percentage_value == "100.0%"
        
        for col in range(1, len(summary_df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = border
            
            # Format specific columns
            if col in [4, 5, 7, 8]:  # Numeric columns
                cell.alignment = Alignment(horizontal="center")
            elif col in [6, 9]:  # Percentage columns
                cell.alignment = Alignment(horizontal="center")
                if col == 9:
                    cell.number_format = '0.0"%"'
            elif col == 10:  # Pendiente (Q)
                if cell.value and cell.value != "" and cell.value != "Q0.00":
                    cell.font = po_font
                cell.alignment = Alignment(horizontal="right")
            elif col == 12:  # Enlace Proyecto
                url = summary_df.iloc[row-2]['Enlace Proyecto']
                cell.value = "Abrir"
                cell.hyperlink = url
                cell.font = link_font
                cell.alignment = Alignment(horizontal="center")
            
            if is_complete and col == 6:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell.font = Font(bold=True, color="006100")
                cell.alignment = Alignment(horizontal="center")
    
    # Set column widths
    column_widths = {
        'A': 35,  # Nombre del Proyecto
        'B': 12,  # Fecha Inicio
        'C': 12,  # Fecha Fin
        'D': 12,  # Total Tareas
        'E': 18,  # Tareas Completadas
        'F': 14,  # % Completado
        'G': 18,  # Horas Planificadas
        'H': 14,  # Horas Reales
        'I': 18,  # Progreso Promedio
        'J': 18,  # Pendiente (Q)
        'K': 45,  # Órdenes de Compra
        'L': 12   # Enlace Proyecto
    }
    
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width
    
    # Format dates
    for row in range(2, len(summary_df) + 2):
        for col in [2, 3]:  # Fecha Inicio and Fecha Fin
            cell = worksheet.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, datetime):
                cell.number_format = 'YYYY-MM-DD'
    
    worksheet.freeze_panes = 'B2'
    
    # Add totals row
    totals_row = len(summary_df) + 2
    for col in range(1, 13):
        cell = worksheet.cell(row=totals_row - 1, column=col)
        cell.border = Border(bottom=Side(style='double'))
    
    worksheet.cell(row=totals_row, column=1, value="TOTALES:").font = Font(bold=True)
    worksheet.cell(row=totals_row, column=4, value=f"=SUM(D2:D{len(summary_df)+1})").font = Font(bold=True)
    worksheet.cell(row=totals_row, column=5, value=f"=SUM(E2:E{len(summary_df)+1})").font = Font(bold=True)
    worksheet.cell(row=totals_row, column=7, value=f"=SUM(G2:G{len(summary_df)+1})").font = Font(bold=True)
    worksheet.cell(row=totals_row, column=8, value=f"=SUM(H2:H{len(summary_df)+1})").font = Font(bold=True)
    
    # Average progress
    avg_progress_cell = worksheet.cell(row=totals_row, column=9)
    avg_progress_cell.value = f"=AVERAGE(I2:I{len(summary_df)+1})"
    avg_progress_cell.number_format = '0.0"%"'
    avg_progress_cell.font = Font(bold=True)
    avg_progress_cell.alignment = Alignment(horizontal="center")
    
    # Calculate overall percentage
    percentage_cell = worksheet.cell(row=totals_row, column=6)
    percentage_cell.value = f"=IF(D{totals_row}>0,E{totals_row}/D{totals_row},0)"
    percentage_cell.number_format = '0.0%'
    percentage_cell.font = Font(bold=True)
    percentage_cell.alignment = Alignment(horizontal="center")
    
    # Apply borders to totals row
    totals_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    for col in range(1, 13):
        cell = worksheet.cell(row=totals_row, column=col)
        cell.border = border
        cell.fill = totals_fill
        if col in [4, 5, 6, 7, 8, 9]:
            cell.alignment = Alignment(horizontal="center")
    
    # Add legend
    legend_row = totals_row + 3
    add_color_legend(writer, 'Resumen de Proyectos', legend_row, include_summary_colors=True)

def main():
    print("=" * 80)
    print("ODOO PROJECT PLAN EXTRACTION - WITH CURRENCY CONVERSION & TASK SORTING")
    print("=" * 80)
    print(f"Extraction started at: {datetime.now()}")
    print(f"Currency conversion: USD to Q at rate {USD_TO_GTQ_RATE}")
    
    try:
        # Fetch project data
        print("\nFetching project plan data with PO information and currency details...")
        df = fetch_project_data(query_project_plan, db_params)
        
        # Create hierarchical structures with sorting
        print("Creating project hierarchy with Torelo links, PO data, and task sorting...")
        df_hierarchy = create_project_hierarchy(df)  # All projects
        df_open_hierarchy = create_open_projects_hierarchy(df)  # Open projects only
        
        # Create directory with today's date
        date_folder = datetime.now().strftime('%Y-%m-%d')
        base_dir = BASE_DIR
        full_output_dir = os.path.join(base_dir, date_folder)
        os.makedirs(full_output_dir, exist_ok=True)
        
        # Create filename - KEEP ORIGINAL NAME FORMAT
        filename = f"Odoo_Project_Plan_MS_Project_Style_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        full_output_path = os.path.join(full_output_dir, filename)
        
        # Save to Excel with ALL sheets
        with pd.ExcelWriter(full_output_path, engine="openpyxl") as writer:
            # Open Projects sheet (active projects only)
            format_project_plan_excel(writer, df_open_hierarchy, sheet_name='Proyectos Abiertos', add_status_colors=True)
            add_color_legend(writer, 'Proyectos Abiertos', len(df_open_hierarchy) + 4, include_summary_colors=False)
            
            # All Projects sheet (complete hierarchical view)
            format_project_plan_excel(writer, df_hierarchy, sheet_name='Todos los Proyectos', add_status_colors=True)
            add_color_legend(writer, 'Todos los Proyectos', len(df_hierarchy) + 4, include_summary_colors=False)
            
            # Gantt Data sheet - RESTORED
            create_gantt_data_sheet(writer, df)
            
            # Summary sheet with PO information
            create_project_summary_sheet(writer, df)
        
        print(f"\nProject plan saved to: {full_output_path}")
        
        # Print summary
        print("\n" + "=" * 50)
        print("EXTRACTION SUMMARY")
        print("=" * 50)
        print(f"Total Projects: {df['project_id'].nunique()}")
        print(f"Total Tasks: {len(df)}")
        
        # PO Summary with currency breakdown
        total_po_amount = df['total_po_amount'].sum()
        tasks_with_po = len(df[df['po_count'] > 0])
        unique_pos = df[df['po_references'].notna()]['po_references'].str.split(', ', expand=True).stack().nunique() if any(df['po_references'].notna()) else 0
        
        # Check for USD conversions
        usd_tasks = df[df['currencies_used'].notna() & df['currencies_used'].str.contains('USD', na=False)] if 'currencies_used' in df.columns else pd.DataFrame()
        
        print(f"\nPurchase Order Summary:")
        print(f"Tasks with POs: {tasks_with_po}")
        print(f"Unique PO Numbers: {unique_pos}")
        print(f"Total Budget Amount: Q{total_po_amount:,.2f}")
        if len(usd_tasks) > 0:
            print(f"  (includes {len(usd_tasks)} tasks with USD converted at {USD_TO_GTQ_RATE})")
        
        # Task summary
        completed_statuses = ['Hecho', 'hecho', 'Completo', 'completo', 'Done', 'done', 
                             'Completado', 'completado', 'Finalizado', 'finalizado']
        df_open = df[~df['status'].isin(completed_statuses)]
        df_completed = df[df['status'].isin(completed_statuses)]
        
        print(f"\nTask Status:")
        print(f"Open Tasks: {len(df_open)}")
        print(f"Completed Tasks: {len(df_completed)}")
        completion_rate = (len(df_completed)/len(df)*100) if len(df) > 0 else 0
        print(f"Overall Completion Rate: {completion_rate:.1f}%")
        
        # Tasks without fecha_inicio
        tasks_without_start = len(df[df['task_start_date'].isna()])
        print(f"\nTasks without Fecha Inicio: {tasks_without_start}")
        
        # Calculate overdue and due soon
        today = pd.Timestamp.now()
        for idx, row in df.iterrows():
            if pd.notna(row['task_end_date']):
                df.at[idx, 'days_until_due'] = (pd.to_datetime(row['task_end_date']) - today).days
            else:
                df.at[idx, 'days_until_due'] = None
        
        overdue_tasks = 0
        due_soon = 0
        for idx, row in df.iterrows():
            if pd.notna(row['days_until_due']) and not pd.isna(row['days_until_due']) and row['status'] not in completed_statuses:
                try:
                    days = float(row['days_until_due'])
                    if days < 0:
                        overdue_tasks += 1
                    elif 0 <= days <= 3:
                        due_soon += 1
                except:
                    pass
        
        blocked_tasks = len(df[df['status'].str.lower().str.contains('bloqueado|blocked', na=False)])
        
        print(f"\nOverdue Tasks: {overdue_tasks}")
        print(f"Tasks Due Within 3 Days: {due_soon}")
        print(f"Blocked Tasks: {blocked_tasks}")
        print(f"Tasks with Dependencies: {len(df[df['dependencies'].notna()])}")
        print(f"Tasks without End Date: {len(df[df['task_end_date'].isna()])}")
        
        print("\n✅ NEW FEATURES IMPLEMENTED:")
        print("   ✓ Currency conversion: USD → Q at 7.7 exchange rate")
        print("   ✓ All amounts displayed in Quetzales (Q)")
        print("   ✓ USD amounts shown in parentheses (e.g., 'RFQ-P01184 ($22,000)')")
        print("   ✓ Tasks sorted by end date (earliest first)")
        print("   ✓ Projects sorted by their earliest task end date")
        print("   ✓ Hierarchy maintained (subtasks stay under parent tasks)")
        print("   ✓ 'Pendiente (Q)' column replaces 'Total PO Amount'")
        print("   ✓ 'PO/RFQ Numbers' column shows order references")
        
        print("\n📊 Excel Sheets Created:")
        print("   1. Proyectos Abiertos - Open projects sorted by timeline")
        print("   2. Todos los Proyectos - All projects sorted by timeline")
        print("   3. Datos Gantt - Gantt chart data with budget in Q")
        print("   4. Resumen de Proyectos - Summary with totals in Q")
        
        print(f"\nExtraction completed at: {datetime.now()}")
        print("=" * 80)
        
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()