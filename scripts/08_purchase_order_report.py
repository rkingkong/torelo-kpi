#!/usr/bin/env python3
"""
Purchase Order Report
Generates comprehensive PO report with pending payments and receipts
"""
import sys
sys.path.append('/opt/torelo-kpi/scripts')
from config import DATABASE_CONFIG, BASE_DIR, WEB_DIR
import pandas as pd
from sqlalchemy import create_engine
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import datetime
import os
import re
import html

def clean_html_text(text):
    """Remove HTML tags and clean text"""
    if not text or text == '':
        return ''
    
    # Convert to string
    text = str(text)
    
    # Decode HTML entities
    text = html.unescape(text)
    
    # Remove HTML tags
    text = re.sub(r'<br\s*/?>', '\n', text, flags=re.IGNORECASE)
    text = re.sub(r'</p>', '\n', text, flags=re.IGNORECASE)
    text = re.sub(r'<[^>]+>', '', text)
    
    # Clean up extra whitespace
    text = re.sub(r'\n\s*\n', '\n', text)
    text = re.sub(r'^\s+|\s+$', '', text)
    text = re.sub(r' +', ' ', text)
    
    return text.strip()

def clean_excel_value(value):
    """Clean values to remove illegal characters for Excel"""
    if value is None:
        return ""
    
    # Convert to string
    value = str(value)
    
    # Remove illegal characters
    illegal_chars = [chr(i) for i in range(0, 32) if i not in [9, 10, 13]]
    for char in illegal_chars:
        value = value.replace(char, '')
    
    value = value.replace('\x00', '')
    value = value.replace('\r\n', '\n')
    value = value.replace('\r', '\n')
    
    # Clean whitespace
    lines = value.split('\n')
    lines = [' '.join(line.split()) for line in lines]
    value = '\n'.join(line for line in lines if line)
    
    return value

def get_purchase_order_data():
    """Get all purchase order data from Odoo database"""
    
    query = """
    WITH invoice_data AS (
        -- Get invoice data linked to purchase orders
        SELECT 
            pol.order_id,
            pol.id AS line_id,
            am.id AS invoice_id,
            am.name AS invoice_number,
            am.state AS invoice_state,
            am.payment_state,
            am.amount_total_signed AS invoice_total,
            am.amount_residual_signed AS amount_unpaid,
            am.invoice_date,
            am.invoice_date_due
        FROM account_move am
        JOIN account_move_line aml ON aml.move_id = am.id
        JOIN purchase_order_line pol ON pol.id = aml.purchase_line_id
        WHERE am.move_type = 'in_invoice'
        AND am.state != 'cancel'
    )
    SELECT 
        -- Purchase Order Info
        po.id AS po_id,
        po.name AS po_number,
        po.partner_ref AS vendor_reference,
        po.state AS po_state,
        po.date_order::date AS order_date,
        po.date_planned::date AS planned_date,
        po.amount_total AS po_total,
        COALESCE(po.internal_notes, '') AS internal_notes,
        
        -- State display
        CASE 
            WHEN po.state = 'draft' THEN 'Borrador'
            WHEN po.state = 'sent' THEN 'Solicitud Enviada'
            WHEN po.state = 'to approve' THEN 'Por Aprobar'
            WHEN po.state = 'purchase' THEN 'Orden de Compra'
            WHEN po.state = 'done' THEN 'Bloqueado'
            WHEN po.state = 'cancel' THEN 'Cancelado'
            ELSE po.state
        END AS state_display,
        
        -- Supplier Info
        rp.name AS supplier_name,
        rp.phone AS supplier_phone,
        rp.email AS supplier_email,
        
        -- Line Details
        pol.id AS line_id,
        pol.name AS line_description,
        pol.product_qty AS qty_ordered,
        pol.qty_received,
        pol.qty_invoiced,
        pol.price_unit,
        pol.price_total AS line_total,
        pol.date_planned::date AS delivery_date,
        
        -- Product Info
        pt.name AS product_name,
        pp.default_code AS product_code,
        
        -- Invoice Info
        COALESCE(inv.invoice_number, '') AS invoice_number,
        COALESCE(inv.payment_state, 'not_paid') AS payment_state,
        COALESCE(inv.amount_unpaid, 0) AS invoice_unpaid,
        
        -- Currency
        CASE 
            WHEN po.currency_id = 163 THEN 'GTQ'
            WHEN po.currency_id = 2 THEN 'USD'
            ELSE rc.name
        END AS currency,
        
        -- User/Buyer (keeping in query but not in Excel)
        rpu.name AS buyer_name,
        
        -- Calculated Fields
        CASE 
            WHEN pol.qty_invoiced = 0 OR pol.qty_invoiced IS NULL THEN pol.price_total
            WHEN pol.qty_invoiced < pol.product_qty THEN 
                (pol.product_qty - pol.qty_invoiced) * pol.price_unit
            ELSE COALESCE(inv.amount_unpaid, 0)
        END AS amount_pending,
        
        CASE 
            WHEN COALESCE(inv.invoice_date_due, pol.date_planned) < CURRENT_DATE 
            THEN CURRENT_DATE - COALESCE(inv.invoice_date_due, pol.date_planned)::date 
            ELSE 0 
        END AS days_overdue,
        
        -- Reception Status
        pol.product_qty - pol.qty_received AS qty_pending_receipt,
        
        CASE 
            WHEN pol.qty_received = 0 THEN 'Not Received'
            WHEN pol.qty_received < pol.product_qty THEN 'Partially Received'
            ELSE 'Fully Received'
        END AS reception_status,
        
        CASE 
            WHEN pol.qty_invoiced = 0 OR pol.qty_invoiced IS NULL THEN 'Not Invoiced'
            WHEN pol.qty_invoiced < pol.product_qty THEN 'Partially Invoiced'
            WHEN inv.amount_unpaid > 0 THEN 'Invoiced - Unpaid'
            ELSE 'Paid'
        END AS payment_status
        
    FROM purchase_order po
    JOIN res_partner rp ON po.partner_id = rp.id
    JOIN purchase_order_line pol ON po.id = pol.order_id
    JOIN product_product pp ON pol.product_id = pp.id
    JOIN product_template pt ON pp.product_tmpl_id = pt.id
    LEFT JOIN res_users u ON po.user_id = u.id
    LEFT JOIN res_partner rpu ON u.partner_id = rpu.id
    LEFT JOIN res_currency rc ON po.currency_id = rc.id
    LEFT JOIN invoice_data inv ON inv.line_id = pol.id
    
    WHERE 
        po.state NOT IN ('cancel')  -- Exclude cancelled orders
    ORDER BY 
        po.state,
        po.date_order DESC
    """
    
    # Connect and fetch data
    print("Connecting to database...")
    conn_str = f"postgresql://{DATABASE_CONFIG['user']}:{DATABASE_CONFIG['password']}@{DATABASE_CONFIG['host']}:{DATABASE_CONFIG['port']}/{DATABASE_CONFIG['dbname']}"
    engine = create_engine(conn_str)
    
    try:
        df = pd.read_sql_query(query, engine)
        print(f"Fetched {len(df)} purchase order line records")
        
        # Clean product names
        df['product_name'] = df['product_name'].apply(lambda x: 
            x.get('es_GT', x.get('en_US', str(x))) if isinstance(x, dict) else str(x)
        )
        
        # Clean HTML from internal_notes field first
        df['internal_notes'] = df['internal_notes'].apply(clean_html_text)
        
        # Then clean all text columns for Excel
        text_columns = ['po_number', 'vendor_reference', 'supplier_name', 'supplier_phone', 
                        'supplier_email', 'line_description', 'product_name', 'product_code', 
                        'invoice_number', 'buyer_name', 'internal_notes']
        
        for col in text_columns:
            if col in df.columns:
                df[col] = df[col].apply(clean_excel_value)
        
        return df
    finally:
        engine.dispose()
 
def create_excel_report(df, output_path):
    """Create Excel report with 3 tabs"""
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Color scheme
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    overdue_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    pending_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Define truly open states (exclude 'done' which is 'bloqueado')
    open_states = ['draft', 'sent', 'to approve', 'purchase']
    
    # Split data by state
    df_open = df[df['po_state'].isin(open_states)].copy()
    df_done = df[df['po_state'] == 'done'].copy()
    
    # 1. OPEN POs TAB - AGGREGATED BY PO (WITHOUT BUYER)
    ws_po_summary = wb.create_sheet("Open POs Summary")
    
    # Aggregate by PO - include notes but EXCLUDE buyer_name
    po_summary = df_open[df_open['amount_pending'] > 0].groupby([
        'po_number', 'vendor_reference', 'supplier_name', 'currency', 
        'order_date', 'state_display', 'internal_notes'
    ]).agg({
        'amount_pending': 'sum',
        'line_total': 'sum'
    }).reset_index()
    
    # Calculate Amount Paid column
    po_summary['amount_paid'] = po_summary['line_total'] - po_summary['amount_pending']
    
    po_summary.columns = ['PO Number', 'Vendor Reference', 'Supplier', 'Currency', 
                          'Order Date', 'Status', 'Internal Notes', 
                          'Amount Pending', 'Total Amount', 'Amount Paid']
    
    # Sort by amount pending
    po_summary = po_summary.sort_values('Amount Pending', ascending=False)
    
    # Write headers
    for c_idx, col in enumerate(po_summary.columns, 1):
        cell = ws_po_summary.cell(row=1, column=c_idx, value=col)
        cell.fill = header_fill
        cell.font = header_font
    
    # Write data
    for r_idx, row in po_summary.iterrows():
        for c_idx, (col_name, value) in enumerate(row.items(), 1):
            cell = ws_po_summary.cell(row=r_idx+2, column=c_idx, value=value)
            
            # Format cells
            if col_name in ['Amount Pending', 'Total Amount', 'Amount Paid']:
                cell.number_format = '#,##0.00'
            elif col_name == 'Order Date':
                if pd.notna(value):
                    cell.number_format = 'dd/mm/yyyy'
            
            # Wrap text for notes column
            if col_name == 'Internal Notes':
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # 2. PENDING RECEIPT LINES TAB - ONLY FROM TRULY OPEN POs (WITHOUT BUYER)
    ws_pending_receipt = wb.create_sheet("Pending Receipt Lines")
    
    # Filter lines pending receipt ONLY from open orders (not done/bloqueado)
    df_pending_receipt = df_open[
        (df_open['qty_pending_receipt'] > 0) & 
        (df_open['po_state'].isin(open_states))  # Double-check state
    ].copy()
    
    # Sort by delivery date
    df_pending_receipt = df_pending_receipt.sort_values(['delivery_date', 'po_number'])
    
    # REMOVED buyer_name from the columns list
    receipt_cols = ['po_number', 'vendor_reference', 'supplier_name', 'product_name', 
                    'product_code', 'line_description', 'qty_ordered', 'qty_received', 
                    'qty_pending_receipt', 'delivery_date', 'days_overdue', 
                    'reception_status', 'state_display', 'internal_notes']
    
    # Write headers
    for c_idx, col in enumerate(receipt_cols, 1):
        cell = ws_pending_receipt.cell(row=1, column=c_idx, value=col.replace('_', ' ').title())
        cell.fill = header_fill
        cell.font = header_font
    
    # Write data
    row_num = 2
    for _, row in df_pending_receipt.iterrows():
        for c_idx, col in enumerate(receipt_cols, 1):
            value = row[col] if col in row else ''
            
            cell = ws_pending_receipt.cell(row=row_num, column=c_idx, value=value)
            
            # Highlight based on status
            if row['days_overdue'] > 0:
                cell.fill = overdue_fill
            elif row['reception_status'] == 'Not Received':
                cell.fill = pending_fill
            
            # Format cells
            if col == 'delivery_date':
                if pd.notna(row[col]):
                    cell.number_format = 'dd/mm/yyyy'
            
            # Wrap text for notes column
            if col == 'internal_notes':
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        row_num += 1
    
    # 3. COMPLETED ORDERS TAB (WITHOUT BUYER)
    ws_done = wb.create_sheet("Done POs - History")
    
    # Aggregate done orders by PO - include notes but EXCLUDE buyer
    done_summary = df_done.groupby([
        'po_number', 'vendor_reference', 'supplier_name', 'currency', 
        'order_date', 'internal_notes'
    ]).agg({
        'line_total': 'sum'
    }).reset_index()
    
    done_summary.columns = ['PO Number', 'Vendor Reference', 'Supplier', 'Currency', 
                           'Order Date', 'Internal Notes', 'Total Amount']
    
    # Write headers
    for c_idx, col in enumerate(done_summary.columns, 1):
        cell = ws_done.cell(row=1, column=c_idx, value=col)
        cell.fill = header_fill
        cell.font = header_font
    
    # Write data
    for r_idx, row in done_summary.iterrows():
        for c_idx, (col_name, value) in enumerate(row.items(), 1):
            cell = ws_done.cell(row=r_idx+2, column=c_idx, value=value)
            
            # Format cells
            if col_name == 'Total Amount':
                cell.number_format = '#,##0.00'
            elif col_name == 'Order Date':
                if pd.notna(value):
                    cell.number_format = 'dd/mm/yyyy'
            
            # Wrap text for notes column
            if col_name == 'Internal Notes':
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Auto-adjust column widths for all sheets
    for ws in wb.worksheets:
        # Add filters
        if ws.max_row > 1:
            ws.auto_filter.ref = f"A1:{chr(64 + len(ws[1])) + str(ws.max_row)}"
        
        # Adjust column widths
        for col_idx in range(1, len(ws[1]) + 1):
            max_length = 0
            col_letter = chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)
            
            # Find the header name for this column
            header_cell = ws.cell(row=1, column=col_idx)
            col_name = header_cell.value if header_cell.value else ""
            
            # Special handling for Internal Notes column - make it wider
            if 'Internal Notes' in str(col_name):
                ws.column_dimensions[col_letter].width = 50
                # Also set row height for notes
                for row in range(2, ws.max_row + 1):
                    if ws.cell(row=row, column=col_idx).value:
                        # Count line breaks to estimate height
                        text = str(ws.cell(row=row, column=col_idx).value)
                        lines = text.count('\n') + 1
                        if lines > 1:
                            ws.row_dimensions[row].height = 15 * lines
            else:
                for row_idx in range(1, min(ws.max_row + 1, 100)):  # Check first 100 rows
                    try:
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width
    
    # Save workbook
    wb.save(output_path)
    print(f"\nReport saved to: {output_path}")

def print_summary(df):
    """Print summary to console"""
    print("\n" + "="*60)
    print("PURCHASE ORDER SUMMARY")
    print("="*60)
    
    # Define open states
    open_states = ['draft', 'sent', 'to approve', 'purchase']
    
    # Split by state
    df_open = df[df['po_state'].isin(open_states)]
    df_done = df[df['po_state'] == 'done']
    
    # Open orders with pending payments
    df_pending_payment = df_open[df_open['amount_pending'] > 0]
    
    print(f"\nOPEN PURCHASE ORDERS:")
    print(f"Total Open Orders: {df_open['po_number'].nunique()}")
    print(f"Orders with Pending Payments: {df_pending_payment['po_number'].nunique()}")
    
    # By currency
    for currency in df_pending_payment['currency'].unique():
        if pd.notna(currency):
            total = df_pending_payment[df_pending_payment['currency'] == currency]['amount_pending'].sum()
            symbol = '$' if currency == 'USD' else 'Q' if currency == 'GTQ' else currency
            print(f"  - Pending {currency}: {symbol}{total:,.2f}")
    
    # Pending receipts (only from open orders)
    df_pending_receipt = df_open[
        (df_open['qty_pending_receipt'] > 0) & 
        (df_open['po_state'].isin(open_states))
    ]
    print(f"\nPENDING RECEIPTS (Open Orders Only):")
    print(f"Lines Pending Receipt: {len(df_pending_receipt)}")
    print(f"Total Qty Pending: {df_pending_receipt['qty_pending_receipt'].sum():,.0f}")
    
    # Orders with notes
    orders_with_notes = df[df['internal_notes'].str.len() > 0]['po_number'].nunique()
    print(f"\nOrders with Internal Notes: {orders_with_notes}")
    
    print(f"\nCOMPLETED ORDERS (Bloqueado):")
    print(f"Total Done Orders: {df_done['po_number'].nunique()}")
    
    # Overdue summary
    overdue = df_pending_payment[df_pending_payment['days_overdue'] > 0]
    if not overdue.empty:
        print(f"\nOVERDUE ITEMS:")
        print(f"Total Overdue: {len(overdue)}")
        print(f"Average Days Overdue: {overdue['days_overdue'].mean():.1f}")
    
    print("\n" + "="*60)

def main():
    """Main execution"""
    print("PURCHASE ORDER REPORT")
    print(f"Started at: {datetime.datetime.now()}")
    print("-"*60)
    
    try:
        # Create date-based output directory
        date_folder = datetime.datetime.now().strftime('%Y-%m-%d')
        output_dir = os.path.join(BASE_DIR, date_folder)
        os.makedirs(output_dir, exist_ok=True)
        
        # Generate filename with timestamp
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Purchase_Order_Report_{timestamp}.xlsx'
        output_path = os.path.join(output_dir, filename)
        
        # Get data
        df = get_purchase_order_data()
        
        if df.empty:
            print("No purchase orders found.")
            return
        
        # Print summary
        print_summary(df)
        
        # Create Excel report
        print("\nGenerating Excel report...")
        create_excel_report(df, output_path)
        
        print("\n✓ COMPLETE! Report generated successfully.")
        print(f"\nThe report includes 3 tabs:")
        print("  1. Open POs Summary - With INTERNAL NOTES and Amount Paid")
        print("  2. Pending Receipt Lines - With INTERNAL NOTES") 
        print("  3. Done POs History - With INTERNAL NOTES")
        print("\nHTML tags have been removed from internal notes for better readability.")
        
    except Exception as e:
        print(f"\n✗ ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()